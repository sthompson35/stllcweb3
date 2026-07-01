"""Dynasty Accounting OS v1.0 - Part 4: REAL ESTATE sheets 50-57"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAV="1A2744";GOLD="D4AF37";DARK="2C3E50";WHITE="FFFFFF";ALT="F8F9FA"
BLUE_TEXT="0000FF";BLACK="000000";GREEN_TEXT="008000";RED_TEXT="FF0000"
TAB_RE="2E7D32"

wb=load_workbook('C:/stllcweb3/Dynasty_Accounting_OS_v1.xlsx')

def hdr_font(size=10,bold=True,color=GOLD): return Font(name='Arial',size=size,bold=bold,color=color)
def data_font(size=10,color=BLACK): return Font(name='Arial',size=size,color=color)
def blue_font(): return Font(name='Arial',size=10,color=BLUE_TEXT)
def bold_font(size=10,color=BLACK): return Font(name='Arial',size=size,bold=True,color=color)
def nav_fill(): return PatternFill('solid',fgColor=NAV)
def dark_fill(): return PatternFill('solid',fgColor=DARK)
def alt_fill(): return PatternFill('solid',fgColor=ALT)
def white_fill(): return PatternFill('solid',fgColor=WHITE)
def center(): return Alignment(horizontal='center',vertical='center',wrap_text=True)
def left(): return Alignment(horizontal='left',vertical='center',wrap_text=True)
def right_align(): return Alignment(horizontal='right',vertical='center')
def thin_border():
    s=Side(style='thin',color='CCCCCC')
    return Border(left=s,right=s,top=s,bottom=s)
def make_title(ws,title,span='A1:F1',size=13):
    ws.merge_cells(span)
    c=ws[span.split(':')[0]]
    c.value=title;c.font=Font(name='Arial',size=size,bold=True,color=GOLD)
    c.fill=nav_fill();c.alignment=center();ws.row_dimensions[1].height=28
def make_header_row(ws,row,headers,col_start=1):
    for i,h in enumerate(headers):
        c=ws.cell(row=row,column=col_start+i,value=h)
        c.font=hdr_font();c.fill=nav_fill();c.alignment=center();c.border=thin_border()
def set_col_widths(ws,widths):
    for col,w in widths.items(): ws.column_dimensions[col].width=w
def style_row(ws,row,ncols,alt=False,col_start=1):
    fill=alt_fill() if alt else white_fill()
    for i in range(ncols):
        c=ws.cell(row=row,column=col_start+i)
        c.fill=fill;c.border=thin_border()

# ============================================================
# SHEET 50: PROPERTY_MASTER
# ============================================================
ws=wb.create_sheet("50_PROPERTY_MASTER")
ws.sheet_properties.tabColor=TAB_RE
make_title(ws,"PROPERTY MASTER — Real Estate Portfolio",'A1:Y1')
ws.row_dimensions[2].height=6

hdrs=['Property_ID','Address','City','State','Zip','Property_Type','Status','Purchase_Date','Purchase_Price','Land_Value','Building_Value','Closing_Costs','ARV','Rehab_Budget','Rehab_Actual','Rehab_Variance','Loan_Balance','Loan_Type','Monthly_Payment','Monthly_Rent','NOI','DSCR','Cap_Rate','Equity','Exit_Strategy']
make_header_row(ws,3,hdrs)

# 502 Buckley data
row=4
prop_inputs=[
    ('A',4,'P001'),('B',4,'502 Buckley'),('C',4,'St. Louis'),('D',4,'MO'),('E',4,'63137'),
    ('F',4,'Single Family'),('G',4,'Active — Rehab'),('H',4,'05/01/2026'),
    ('I',4,114000),('J',4,16100),('K',4,97900),('L',4,4500),
    ('M',4,200000),('N',4,90000),('O',4,11500),
]
for col,r,val in prop_inputs:
    c=ws[f'{col}{r}']
    c.value=val
    if col in ['I','J','K','L','M','N','O']:
        c.font=blue_font();c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
    elif col in ['F','G']:
        c.font=blue_font()
    else:
        c.font=blue_font()

# Formula columns
ws['P4'].value='=N4-O4';ws['P4'].font=Font(name='Arial',size=10,color=BLACK)
ws['P4'].number_format='$#,##0;($#,##0);"-"';ws['P4'].alignment=right_align()

ws['Q4'].value=114000;ws['Q4'].font=blue_font()
ws['Q4'].number_format='$#,##0;($#,##0);"-"';ws['Q4'].alignment=right_align()
ws['R4'].value='FHA';ws['R4'].font=blue_font()
ws['S4'].value=750;ws['S4'].font=blue_font()
ws['S4'].number_format='$#,##0;($#,##0);"-"';ws['S4'].alignment=right_align()
ws['T4'].value=0;ws['T4'].font=blue_font()
ws['T4'].number_format='$#,##0;($#,##0);"-"';ws['T4'].alignment=right_align()

# NOI = (Monthly Rent - OpEx) * 12; simplified: rent * 0.65 * 12
ws['U4'].value='=IF(T4>0,T4*0.65*12,0)'
ws['U4'].font=Font(name='Arial',size=10,color=BLACK)
ws['U4'].number_format='$#,##0;($#,##0);"-"';ws['U4'].alignment=right_align()

# DSCR = NOI / (Monthly Payment * 12)
ws['V4'].value='=IF(S4*12>0,U4/(S4*12),0)'
ws['V4'].font=Font(name='Arial',size=10,color=BLACK)
ws['V4'].number_format='0.00x';ws['V4'].alignment=right_align()

# Cap Rate = NOI / ARV
ws['W4'].value='=IF(M4>0,U4/M4,0)'
ws['W4'].font=Font(name='Arial',size=10,color=BLACK)
ws['W4'].number_format='0.0%';ws['W4'].alignment=right_align()

# Equity = ARV - Loan Balance
ws['X4'].value='=M4-Q4'
ws['X4'].font=Font(name='Arial',size=10,color=BLACK)
ws['X4'].number_format='$#,##0;($#,##0);"-"';ws['X4'].alignment=right_align()

ws['Y4'].value='BRRR';ws['Y4'].font=blue_font()

# Notes row below
ws.cell(row=5,column=1,value='Notes:').font=bold_font()
ws.merge_cells('B5:Y5')
ws.cell(row=5,column=2,value='502 Buckley objectives: Fix crawlspace/floor issues, add basement, add 3rd bedroom. FHA purchase 05/01/2026. Rehab in progress.').font=data_font()

style_row(ws,4,25,alt=False)
style_row(ws,5,25,alt=True)

set_col_widths(ws,{'A':12,'B':22,'C':14,'D':6,'E':8,'F':16,'G':18,'H':14,'I':16,'J':14,'K':16,'L':14,'M':16,'N':14,'O':14,'P':16,'Q':14,'R':10,'S':16,'T':14,'U':14,'V':10,'W':10,'X':14,'Y':14})
ws.freeze_panes='A4'

# ============================================================
# SHEET 51: ACQUISITIONS
# ============================================================
ws=wb.create_sheet("51_ACQUISITIONS")
ws.sheet_properties.tabColor=TAB_RE
make_title(ws,"ACQUISITIONS — Deal Pipeline",'A1:R1')
ws.row_dimensions[2].height=6

hdrs=['Acq_ID','Property','Address','Lead_Source','Lead_Date','Contract_Date','Purchase_Price','ARV','Rehab_Est','MAO','Spread','Under_Contract','Closing_Date','Entity','Deal_Type','Status','Notes']
make_header_row(ws,3,hdrs)

row=4
acq_vals=[('A','ACQ001'),('B','502 Buckley'),('C','502 Buckley, St. Louis MO 63137'),
          ('D','Direct'),('E','04/15/2026'),('F','04/25/2026'),('G',114000),('H',200000),
          ('I',90000),('L','Yes'),('M','05/01/2026'),('N','STLLC'),('O','BRRR'),('P','Closed'),
          ('Q','FHA purchase. Rehab in progress. Objectives: crawlspace, basement, 3rd bedroom')]
for col,val in acq_vals:
    c=ws[f'{col}{row}']
    c.value=val
    if col in ['G','H','I']:
        c.font=blue_font();c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
    else:
        c.font=blue_font()

# MAO formula
ws['J4'].value='=H4*0.7-I4'
ws['J4'].font=Font(name='Arial',size=10,color=BLACK)
ws['J4'].number_format='$#,##0;($#,##0);"-"';ws['J4'].alignment=right_align()

# Spread = MAO - Purchase
ws['K4'].value='=J4-G4'
ws['K4'].font=Font(name='Arial',size=10,color=BLACK)
ws['K4'].number_format='$#,##0;($#,##0);"-"';ws['K4'].alignment=right_align()

style_row(ws,4,17)
set_col_widths(ws,{'A':10,'B':16,'C':32,'D':14,'E':12,'F':14,'G':16,'H':16,'I':14,'J':14,'K':14,'L':14,'M':14,'N':10,'O':12,'P':10,'Q':45})
ws.freeze_panes='A4'

# ============================================================
# SHEET 52: REHAB_ENGINE
# ============================================================
ws=wb.create_sheet("52_REHAB_ENGINE")
ws.sheet_properties.tabColor=TAB_RE
make_title(ws,"REHAB BUDGET & TRACKER — 502 Buckley St. Louis MO",'A1:L1',size=12)
ws.row_dimensions[2].height=6

hdrs=['Work_Order','Category','Scope_of_Work','Budget','Actual','Variance','Pct_Complete','Status','Contractor','Start_Date','End_Date','Notes']
make_header_row(ws,3,hdrs)

rehab_items=[
    ('WO001','Demo & Site Prep','Clear structure, remove debris, dumpsters',5000,3500,70,'In Progress','TBD','05/10/2026','05/31/2026','Phase 1'),
    ('WO002','Crawlspace/Foundation','Fix crawlspace, address foundation issues, add drainage',12000,0,0,'Pending','TBD','','','Critical — must complete first'),
    ('WO003','Basement Addition','Excavate and pour basement, waterproof',20000,0,0,'Pending','TBD','','','Major addition'),
    ('WO004','Framing','Add 3rd bedroom framing, structural repairs',8000,0,0,'Pending','TBD','','',''),
    ('WO005','Roofing','Full roof replacement',6000,0,0,'Pending','TBD','','',''),
    ('WO006','Exterior/Siding','Siding, exterior paint, gutters',4000,0,0,'Pending','TBD','','',''),
    ('WO007','Windows & Doors','Replace all windows and exterior doors',5000,0,0,'Pending','TBD','','',''),
    ('WO008','Electrical','Full electrical upgrade, 200 amp panel',8000,0,0,'Pending','TBD','','',''),
    ('WO009','Plumbing','Rough plumbing, fixtures, water heater',6000,0,0,'Pending','TBD','','',''),
    ('WO010','HVAC','New forced air system, ductwork',7000,0,0,'Pending','TBD','','',''),
    ('WO011','Insulation','Full house insulation',2500,0,0,'Pending','TBD','','',''),
    ('WO012','Drywall','Hang, tape, mud, sand throughout',4000,0,0,'Pending','TBD','','',''),
    ('WO013','Interior Doors','All interior doors and hardware',1500,0,0,'Pending','TBD','','',''),
    ('WO014','Flooring','LVP main floor, carpet bedrooms',5000,0,0,'Pending','TBD','','',''),
    ('WO015','Kitchen','Full kitchen — cabinets, counters, backsplash',8000,0,0,'Pending','TBD','','',''),
    ('WO016','Bathroom 1','Full master bath renovate',4000,0,0,'Pending','TBD','','',''),
    ('WO017','Bathroom 2','Second bathroom add or renovate',3500,0,0,'Pending','TBD','','',''),
    ('WO018','Paint — Interior','Full interior prime + 2 coats',2500,0,0,'Pending','TBD','','',''),
    ('WO019','Paint — Exterior','Exterior paint and trim',1500,0,0,'Pending','TBD','','',''),
    ('WO020','Cabinets/Millwork','Built-ins, closets, trim',2000,0,0,'Pending','TBD','','',''),
    ('WO021','Landscaping','Grading, seed, basic curb appeal',1500,0,0,'Pending','TBD','','',''),
    ('WO022','Driveway/Parking','Concrete or asphalt driveway',2500,0,0,'Pending','TBD','','',''),
    ('WO023','Appliances','Range, fridge, dishwasher, W/D hook',3000,0,0,'Pending','TBD','','',''),
    ('WO024','Lighting/Fixtures','All lighting, ceiling fans, switches',2000,0,0,'Pending','TBD','','',''),
    ('WO025','Cleanup/Final','Final debris removal, touch-ups',1000,0,0,'Pending','TBD','','',''),
    ('WO026','Permits','All building permits',1500,0,0,'Pending','TBD','','',''),
    ('WO027','Inspections','All required inspections',500,0,0,'Pending','TBD','','',''),
]

for i,item in enumerate(rehab_items):
    row=4+i
    wo,cat,scope,budget,actual,pct,status,contr,start,end,notes=item
    ws.cell(row=row,column=1,value=wo).font=data_font()
    ws.cell(row=row,column=2,value=cat).font=bold_font()
    ws.cell(row=row,column=3,value=scope).font=data_font()
    bc=ws.cell(row=row,column=4,value=budget)
    bc.font=blue_font();bc.number_format='$#,##0;($#,##0);"-"';bc.alignment=right_align()
    ac=ws.cell(row=row,column=5,value=actual)
    ac.font=blue_font();ac.number_format='$#,##0;($#,##0);"-"';ac.alignment=right_align()
    vc=ws.cell(row=row,column=6,value=f'=D{row}-E{row}')
    vc.font=Font(name='Arial',size=10,color=BLACK);vc.number_format='$#,##0;($#,##0);"-"';vc.alignment=right_align()
    pc=ws.cell(row=row,column=7,value=pct)
    pc.font=blue_font();pc.number_format='0%';pc.alignment=right_align()
    ws.cell(row=row,column=8,value=status).font=data_font()
    ws.cell(row=row,column=9,value=contr).font=data_font()
    ws.cell(row=row,column=10,value=start).font=data_font()
    ws.cell(row=row,column=11,value=end).font=data_font()
    ws.cell(row=row,column=12,value=notes).font=data_font()
    style_row(ws,row,12,alt=(i%2==1))

# Contingency and GC rows
base_sum_end=4+len(rehab_items)-1
cont_row=4+len(rehab_items)
ws.cell(row=cont_row,column=1,value='WO028').font=data_font()
ws.cell(row=cont_row,column=2,value='Contingency 10%').font=bold_font()
ws.cell(row=cont_row,column=3,value='10% contingency reserve').font=data_font()
cf=ws.cell(row=cont_row,column=4,value=f'=SUM(D4:D{base_sum_end})*0.1')
cf.font=Font(name='Arial',size=10,color=BLACK);cf.number_format='$#,##0;($#,##0);"-"';cf.alignment=right_align()
ws.cell(row=cont_row,column=5,value=0).font=blue_font()
ws.cell(row=cont_row,column=5).number_format='$#,##0;($#,##0);"-"';ws.cell(row=cont_row,column=5).alignment=right_align()
ws.cell(row=cont_row,column=6,value=f'=D{cont_row}-E{cont_row}').font=Font(name='Arial',size=10,color=BLACK)
ws.cell(row=cont_row,column=6).number_format='$#,##0;($#,##0);"-"';ws.cell(row=cont_row,column=6).alignment=right_align()
style_row(ws,cont_row,12,alt=False)

gc_row=cont_row+1
ws.cell(row=gc_row,column=1,value='WO029').font=data_font()
ws.cell(row=gc_row,column=2,value='GC Overhead 10%').font=bold_font()
ws.cell(row=gc_row,column=3,value='General contractor markup').font=data_font()
gf=ws.cell(row=gc_row,column=4,value=f'=SUM(D4:D{base_sum_end})*0.1')
gf.font=Font(name='Arial',size=10,color=BLACK);gf.number_format='$#,##0;($#,##0);"-"';gf.alignment=right_align()
ws.cell(row=gc_row,column=5,value=0).font=blue_font()
ws.cell(row=gc_row,column=5).number_format='$#,##0;($#,##0);"-"';ws.cell(row=gc_row,column=5).alignment=right_align()
ws.cell(row=gc_row,column=6,value=f'=D{gc_row}-E{gc_row}').font=Font(name='Arial',size=10,color=BLACK)
ws.cell(row=gc_row,column=6).number_format='$#,##0;($#,##0);"-"';ws.cell(row=gc_row,column=6).alignment=right_align()
style_row(ws,gc_row,12,alt=True)

co_row=gc_row+1
ws.cell(row=co_row,column=1,value='WO030').font=data_font()
ws.cell(row=co_row,column=2,value='Change Orders').font=bold_font()
ws.cell(row=co_row,column=3,value='Approved change orders').font=data_font()
ws.cell(row=co_row,column=4,value=0).font=blue_font()
ws.cell(row=co_row,column=4).number_format='$#,##0;($#,##0);"-"';ws.cell(row=co_row,column=4).alignment=right_align()
ws.cell(row=co_row,column=5,value=8000).font=blue_font()
ws.cell(row=co_row,column=5).number_format='$#,##0;($#,##0);"-"';ws.cell(row=co_row,column=5).alignment=right_align()
ws.cell(row=co_row,column=6,value=f'=D{co_row}-E{co_row}').font=Font(name='Arial',size=10,color=BLACK)
ws.cell(row=co_row,column=6).number_format='$#,##0;($#,##0);"-"';ws.cell(row=co_row,column=6).alignment=right_align()
style_row(ws,co_row,12,alt=False)

total_row=co_row+2
ws.merge_cells(f'A{total_row}:C{total_row}')
ws.cell(row=total_row,column=1,value='TOTAL REHAB BUDGET').font=bold_font(size=11)
ws.cell(row=total_row,column=1).fill=nav_fill()
ws.cell(row=total_row,column=1).font=Font(name='Arial',size=11,bold=True,color=GOLD)
tc=ws.cell(row=total_row,column=4,value=f'=SUM(D4:D{co_row})')
tc.font=Font(name='Arial',size=11,bold=True,color=GOLD);tc.number_format='$#,##0;($#,##0);"-"'
tc.alignment=right_align();tc.fill=nav_fill();tc.border=thin_border()
ta=ws.cell(row=total_row,column=5,value=f'=SUM(E4:E{co_row})')
ta.font=Font(name='Arial',size=11,bold=True,color=GOLD);ta.number_format='$#,##0;($#,##0);"-"'
ta.alignment=right_align();ta.fill=nav_fill();ta.border=thin_border()
tv=ws.cell(row=total_row,column=6,value=f'=D{total_row}-E{total_row}')
tv.font=Font(name='Arial',size=11,bold=True,color=GOLD);tv.number_format='$#,##0;($#,##0);"-"'
tv.alignment=right_align();tv.fill=nav_fill();tv.border=thin_border()

health_row=total_row+1
ws.merge_cells(f'A{health_row}:C{health_row}')
ws.cell(row=health_row,column=1,value='REHAB HEALTH SCORE:').font=bold_font()
health=ws.cell(row=health_row,column=4,value=f'=IF(E{total_row}>D{total_row},"RED — OVER BUDGET",IF(E{total_row}/D{total_row}>0.9,"YELLOW — AT RISK","GREEN — ON TRACK"))')
health.font=Font(name='Arial',size=10,bold=True,color=BLACK)

set_col_widths(ws,{'A':10,'B':22,'C':38,'D':14,'E':14,'F':14,'G':12,'H':14,'I':16,'J':14,'K':14,'L':32})
ws.freeze_panes='A4'

# ============================================================
# SHEET 53: DRAWS
# ============================================================
ws=wb.create_sheet("53_DRAWS")
ws.sheet_properties.tabColor=TAB_RE
make_title(ws,"CONSTRUCTION DRAW SCHEDULE",'A1:M1')
ws.row_dimensions[2].height=6

hdrs=['Draw_ID','Date','Property','Lender','Draw_Amount','Draw_Purpose','Work_Completed_%','Inspector_Approved','Funded','Draw_Number','Cumulative_Draws','Remaining_Budget','Notes']
make_header_row(ws,3,hdrs)

draw_data=[
    ('DR001','05/15/2026','502 Buckley','FHA/Personal',3500,'Demo phase materials — lumber, dumpster',70,'No','Yes',1,3500,'=90000-K4','Initial materials draw'),
]
for i,row_data in enumerate(draw_data):
    row=4+i
    for j,val in enumerate(row_data):
        c=ws.cell(row=row,column=j+1,value=val)
        if j==4: c.font=blue_font();c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        elif j==10: c.font=Font(name='Arial',size=10,color=BLACK);c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        elif j==11: c.font=Font(name='Arial',size=10,color=BLACK);c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        else: c.font=data_font()
        c.border=thin_border();c.fill=white_fill()

set_col_widths(ws,{'A':10,'B':12,'C':16,'D':16,'E':14,'F':35,'G':16,'H':18,'I':10,'J':14,'K':18,'L':18,'M':28})
ws.freeze_panes='A4'

# ============================================================
# SHEET 54: HOLDING_COSTS
# ============================================================
ws=wb.create_sheet("54_HOLDING_COSTS")
ws.sheet_properties.tabColor=TAB_RE
make_title(ws,"HOLDING COSTS TRACKER",'A1:L1')
ws.row_dimensions[2].height=6

hdrs=['HC_ID','Property','Month','Mortgage','Property_Tax','Insurance','Utilities','HOA','Other','Total_Monthly','Cumulative_Total','Notes']
make_header_row(ws,3,hdrs)

hold_data=[('HC001','502 Buckley','May 2026',750,150,100,80,0,0)]
for i,row_data in enumerate(hold_data):
    row=4+i
    for j,val in enumerate(row_data):
        c=ws.cell(row=row,column=j+1,value=val)
        if j in range(3,9): c.font=blue_font();c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        else: c.font=data_font()
        c.border=thin_border();c.fill=white_fill()
    tot=ws.cell(row=row,column=10,value=f'=SUM(D{row}:I{row})')
    tot.font=Font(name='Arial',size=10,color=BLACK);tot.number_format='$#,##0;($#,##0);"-"';tot.alignment=right_align()
    tot.border=thin_border();tot.fill=white_fill()
    cum=ws.cell(row=row,column=11,value=f'=J{row}')
    cum.font=Font(name='Arial',size=10,color=BLACK);cum.number_format='$#,##0;($#,##0);"-"';cum.alignment=right_align()
    cum.border=thin_border();cum.fill=white_fill()
    ws.cell(row=row,column=12).border=thin_border();ws.cell(row=row,column=12).fill=white_fill()

set_col_widths(ws,{'A':10,'B':16,'C':12,'D':14,'E':14,'F':12,'G':12,'H':10,'I':10,'J':16,'K':18,'L':28})
ws.freeze_panes='A4'

# ============================================================
# SHEET 55: RENTS
# ============================================================
ws=wb.create_sheet("55_RENTS")
ws.sheet_properties.tabColor=TAB_RE
make_title(ws,"RENT ROLL — Rental Income Tracker",'A1:X1')
ws.row_dimensions[2].height=6

months=['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
hdrs=['Rent_ID','Property','Unit','Tenant_Name','Lease_Start','Lease_End','Monthly_Rent','Security_Deposit','Deposit_Held']+months+['Annual_Rent','Collected_YTD','Notes']
make_header_row(ws,3,hdrs)

# 502 Buckley vacant
row=4
ws.cell(row=row,column=1,value='RENT001').font=data_font()
ws.cell(row=row,column=2,value='502 Buckley').font=data_font()
ws.cell(row=row,column=3,value='Unit 1').font=data_font()
ws.cell(row=row,column=4,value='VACANT').font=Font(name='Arial',size=10,color='FF0000',bold=True)
ws.cell(row=row,column=5,value='').font=data_font()
ws.cell(row=row,column=6,value='').font=data_font()
ws.cell(row=row,column=7,value=1400).font=blue_font()
ws.cell(row=row,column=7).number_format='$#,##0;($#,##0);"-"'
for j in range(9,22): ws.cell(row=row,column=j,value=0).font=data_font()
ar=ws.cell(row=row,column=22,value=f'=SUM(J{row}:U{row})')
ar.font=Font(name='Arial',size=10,color=BLACK);ar.number_format='$#,##0;($#,##0);"-"'
cy=ws.cell(row=row,column=23,value=f'=SUM(J{row}:N{row})')
cy.font=Font(name='Arial',size=10,color=BLACK);cy.number_format='$#,##0;($#,##0);"-"'
ws.cell(row=row,column=24,value='Pending rehab completion — est. rental ready Q4 2026').font=data_font()
style_row(ws,row,24)

set_col_widths(ws,{'A':10,'B':16,'C':8,'D':18,'E':12,'F':12,'G':14,'H':16,'I':14,'J':8,'K':8,'L':8,'M':8,'N':8,'O':8,'P':8,'Q':8,'R':8,'S':8,'T':8,'U':8,'V':8,'W':14,'X':38})
ws.freeze_panes='A4'

# ============================================================
# SHEET 56: CAPEX
# ============================================================
ws=wb.create_sheet("56_CAPEX")
ws.sheet_properties.tabColor=TAB_RE
make_title(ws,"CAPITAL EXPENDITURES — Asset Register",'A1:L1')
ws.row_dimensions[2].height=6

hdrs=['CapEx_ID','Date','Property','Description','Amount','Asset_Class','Useful_Life','Annual_Depreciation','Accumulated_Depr','Book_Value','Capitalized','Notes']
make_header_row(ws,3,hdrs)

capex_data=[
    ('CE001','05/01/2026','502 Buckley','Building purchase',97900,'Residential RE',27.5,'=E4/G4',0,'=E4-I4','Yes','FHA purchase allocation'),
    ('CE002','05/01/2026','502 Buckley','Land purchase',16100,'Land',0,0,0,'=E5',  'Yes','Land not depreciated'),
]
for i,row_data in enumerate(capex_data):
    row=4+i
    capex_id,date,prop,desc,amt,asset_class,life,annual_depr,accum,book,cap,notes=row_data
    ws.cell(row=row,column=1,value=capex_id).font=data_font()
    ws.cell(row=row,column=2,value=date).font=data_font()
    ws.cell(row=row,column=3,value=prop).font=data_font()
    ws.cell(row=row,column=4,value=desc).font=data_font()
    ac=ws.cell(row=row,column=5,value=amt)
    ac.font=blue_font();ac.number_format='$#,##0;($#,##0);"-"';ac.alignment=right_align()
    ws.cell(row=row,column=6,value=asset_class).font=data_font()
    lc=ws.cell(row=row,column=7,value=life)
    lc.font=blue_font();lc.alignment=right_align()
    dc=ws.cell(row=row,column=8,value=annual_depr)
    dc.font=Font(name='Arial',size=10,color=BLACK);dc.number_format='$#,##0;($#,##0);"-"';dc.alignment=right_align()
    xc=ws.cell(row=row,column=9,value=accum)
    xc.font=blue_font();xc.number_format='$#,##0;($#,##0);"-"';xc.alignment=right_align()
    bc=ws.cell(row=row,column=10,value=book)
    bc.font=Font(name='Arial',size=10,color=BLACK);bc.number_format='$#,##0;($#,##0);"-"';bc.alignment=right_align()
    ws.cell(row=row,column=11,value=cap).font=data_font()
    ws.cell(row=row,column=12,value=notes).font=data_font()
    style_row(ws,row,12,alt=(i%2==1))

set_col_widths(ws,{'A':10,'B':12,'C':16,'D':28,'E':14,'F':18,'G':12,'H':20,'I':20,'J':14,'K':12,'L':32})
ws.freeze_panes='A4'

# ============================================================
# SHEET 57: DISPOSITIONS
# ============================================================
ws=wb.create_sheet("57_DISPOSITIONS")
ws.sheet_properties.tabColor=TAB_RE
make_title(ws,"DISPOSITIONS — Property Sales Log",'A1:P1')
ws.row_dimensions[2].height=6

hdrs=['Disp_ID','Property','Sale_Date','Sale_Price','Purchase_Price','Rehab_Cost','Holding_Costs','Selling_Costs','Total_Costs','Net_Profit','ROI','Annualized_ROI','Hold_Days','Exit_Type','Buyer','Notes']
make_header_row(ws,3,hdrs)

ws.cell(row=4,column=1,value='[No dispositions yet]').font=Font(name='Arial',size=10,italic=True,color='888888')
ws.merge_cells('A4:P4')

set_col_widths(ws,{'A':12,'B':18,'C':12,'D':14,'E':16,'F':14,'G':14,'H':14,'I':14,'J':14,'K':10,'L':14,'M':12,'N':12,'O':18,'P':28})
ws.freeze_panes='A4'

wb.save('C:/stllcweb3/Dynasty_Accounting_OS_v1.xlsx')
print("Real Estate sheets (50-57) done")
