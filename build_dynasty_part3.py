"""Dynasty Accounting OS v1.0 - Part 3: BANKING (20-23) and AR/AP (30-43) sheets"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as get_col_letter
from openpyxl.worksheet.datavalidation import DataValidation

NAV="1A2744";GOLD="D4AF37";DARK="2C3E50";WHITE="FFFFFF";ALT="F8F9FA"
BLUE_TEXT="0000FF";BLACK="000000";GREEN_TEXT="008000"
TAB_BANK="008080";TAB_ARAP="FF6600"

wb = load_workbook('C:/stllcweb3/Dynasty_Accounting_OS_v1.xlsx')

def hdr_font(size=10,bold=True,color=GOLD): return Font(name='Arial',size=size,bold=bold,color=color)
def data_font(size=10,color=BLACK): return Font(name='Arial',size=size,color=color)
def blue_font(): return Font(name='Arial',size=10,color=BLUE_TEXT)
def bold_font(size=10,color=BLACK): return Font(name='Arial',size=size,bold=True,color=color)
def nav_fill(): return PatternFill('solid',fgColor=NAV)
def dark_fill(): return PatternFill('solid',fgColor=DARK)
def alt_fill(): return PatternFill('solid',fgColor=ALT)
def white_fill(): return PatternFill('solid',fgColor=WHITE)
def center(): return Alignment(horizontal='center',vertical='center',wrap_text=True)
def left(): return Alignment(horizontal='left',vertical='center',wrap_text=True)
def right_align(): return Alignment(horizontal='right',vertical='center')
def thin_border():
    s=Side(style='thin',color='CCCCCC')
    return Border(left=s,right=s,top=s,bottom=s)
def make_title(ws,title,span='A1:F1',size=13):
    ws.merge_cells(span)
    c=ws[span.split(':')[0]]
    c.value=title;c.font=Font(name='Arial',size=size,bold=True,color=GOLD)
    c.fill=nav_fill();c.alignment=center();ws.row_dimensions[1].height=28
def make_header_row(ws,row,headers,col_start=1):
    for i,h in enumerate(headers):
        c=ws.cell(row=row,column=col_start+i,value=h)
        c.font=hdr_font();c.fill=nav_fill();c.alignment=center();c.border=thin_border()
def set_col_widths(ws,widths):
    for col,w in widths.items(): ws.column_dimensions[col].width=w
def style_row(ws,row,ncols,alt=False,col_start=1):
    fill=alt_fill() if alt else white_fill()
    for i in range(ncols):
        c=ws.cell(row=row,column=col_start+i)
        c.fill=fill;c.border=thin_border()

# ============================================================
# SHEET 20: BANK_ACCOUNTS
# ============================================================
ws=wb.create_sheet("20_BANK_ACCOUNTS")
ws.sheet_properties.tabColor=TAB_BANK
make_title(ws,"BANK ACCOUNTS REGISTRY",'A1:J1')
ws.row_dimensions[2].height=6

hdrs=['Account_ID','Bank_Name','Account_Type','Account_Number','Routing_Number','Opening_Balance','Current_Balance','Last_Reconciled','Status','Notes']
make_header_row(ws,3,hdrs)

bank_data=[
    ('BA001','Chase Bank','Business Checking','XXXX-1234','021000021',0,17200,'05/29/2026','Active','Primary operating account'),
    ('BA002','Chase Bank','Business Savings','XXXX-5678','021000021',0,0,'','Active','Reserve fund'),
]
for i,row_data in enumerate(bank_data):
    row=4+i
    for j,val in enumerate(row_data):
        c=ws.cell(row=row,column=j+1,value=val)
        c.font=blue_font() if j in [6] else data_font()
        if j in [5,6]: c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        c.border=thin_border();c.fill=alt_fill() if i%2==1 else white_fill()

set_col_widths(ws,{'A':12,'B':16,'C':20,'D':16,'E':14,'F':16,'G':16,'H':16,'I':10,'J':35})
ws.freeze_panes='A4'

# ============================================================
# SHEET 21: BANK_TRANSACTIONS
# ============================================================
ws=wb.create_sheet("21_BANK_TRANSACTIONS")
ws.sheet_properties.tabColor=TAB_BANK
make_title(ws,"BANK TRANSACTIONS FEED",'A1:L1')
ws.row_dimensions[2].height=6

hdrs=['Trans_ID','Date','Account_ID','Description','Amount','Trans_Type','Category','GL_Account','Matched','Reconciled','Reference','Notes']
make_header_row(ws,3,hdrs)

trans_data=[
    ('BT001','05/01/2026','BA001','Capital contribution',25000,'Deposit','Capital',3030,'Yes','Yes','WIRE-001','Opening capital'),
    ('BT002','05/01/2026','BA001','502 Buckley closing costs',-4500,'Withdrawal','Acquisition',5010,'Yes','Yes','CHECK-001',''),
    ('BT003','05/15/2026','BA001','Demo materials',-3500,'Withdrawal','Rehab',5110,'Yes','Yes','CHECK-002',''),
    ('BT004','05/25/2026','BA001','PropWire subscription',-299,'Withdrawal','Marketing',6120,'Yes','Yes','CC-001',''),
    ('BT005','05/29/2026','BA001','Misc income',500,'Deposit','Other',4400,'No','No','','Pending match'),
]
for i,row_data in enumerate(trans_data):
    row=4+i
    for j,val in enumerate(row_data):
        c=ws.cell(row=row,column=j+1,value=val)
        c.font=blue_font() if j==4 else data_font()
        if j==4: c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        c.border=thin_border();c.fill=alt_fill() if i%2==1 else white_fill()

set_col_widths(ws,{'A':10,'B':12,'C':12,'D':32,'E':14,'F':14,'G':16,'H':12,'I':10,'J':12,'K':14,'L':28})
ws.freeze_panes='A4'

# ============================================================
# SHEET 22: RECONCILIATION
# ============================================================
ws=wb.create_sheet("22_RECONCILIATION")
ws.sheet_properties.tabColor=TAB_BANK
make_title(ws,"BANK RECONCILIATION — Chase Business Checking — May 2026",'A1:D1')

ws.merge_cells('A3:D3')
c=ws['A3'];c.value="BANK STATEMENT SECTION";c.font=hdr_font();c.fill=dark_fill();c.alignment=left()

rows_data=[
    (4,'Bank Statement Balance','',17200,'$#,##0;($#,##0);"-"',True),
    (5,'+ Deposits in Transit','',0,'$#,##0;($#,##0);"-"',True),
    (6,'- Outstanding Checks','',0,'$#,##0;($#,##0);"-"',True),
    (7,'ADJUSTED BANK BALANCE','=D4+D5-D6',None,'$#,##0;($#,##0);"-"',False),
]
for rrow,label,formula,val,fmt,is_input in rows_data:
    ws.cell(row=rrow,column=1,value=label).font=bold_font() if not is_input else data_font()
    if formula:
        c=ws.cell(row=rrow,column=4,value=formula)
        c.font=bold_font();c.number_format=fmt;c.alignment=right_align()
    elif val is not None:
        c=ws.cell(row=rrow,column=4,value=val)
        c.font=blue_font();c.number_format=fmt;c.alignment=right_align()
    for col in range(1,5):
        ws.cell(row=rrow,column=col).border=thin_border()

ws.merge_cells('A9:D9')
c=ws['A9'];c.value="BOOK (GL) SECTION";c.font=hdr_font();c.fill=dark_fill();c.alignment=left()

book_rows=[
    (10,'Book Balance (per GL)','=SUMIF(\'21_BANK_TRANSACTIONS\'!C:C,"BA001",\'21_BANK_TRANSACTIONS\'!E:E)',None,'$#,##0;($#,##0);"-"',False),
    (11,'+ Interest Earned','',0,'$#,##0;($#,##0);"-"',True),
    (12,'- Bank Charges','',0,'$#,##0;($#,##0);"-"',True),
    (13,'ADJUSTED BOOK BALANCE','=D10+D11-D12',None,'$#,##0;($#,##0);"-"',False),
]
for rrow,label,formula,val,fmt,is_input in book_rows:
    ws.cell(row=rrow,column=1,value=label).font=data_font()
    if formula:
        c=ws.cell(row=rrow,column=4,value=formula)
        c.font=Font(name='Arial',size=10,bold=rrow in [10,13],color=BLACK)
        c.number_format=fmt;c.alignment=right_align()
    elif val is not None:
        c=ws.cell(row=rrow,column=4,value=val)
        c.font=blue_font();c.number_format=fmt;c.alignment=right_align()
    for col in range(1,5):
        ws.cell(row=rrow,column=col).border=thin_border()

ws.cell(row=15,column=1,value='DIFFERENCE (must = $0)').font=bold_font(size=11)
diff=ws.cell(row=15,column=4,value='=D7-D13')
diff.font=bold_font(size=11);diff.number_format='$#,##0;($#,##0);"-"';diff.alignment=right_align()
for col in range(1,5): ws.cell(row=15,column=col).border=thin_border()
ws.cell(row=15,column=4).fill=PatternFill('solid',fgColor='E8F5E9')

set_col_widths(ws,{'A':35,'B':20,'C':20,'D':18})

# ============================================================
# SHEET 23: CREDIT_CARDS
# ============================================================
ws=wb.create_sheet("23_CREDIT_CARDS")
ws.sheet_properties.tabColor=TAB_BANK
make_title(ws,"CREDIT CARD TRANSACTIONS",'A1:K1')
ws.row_dimensions[2].height=6

hdrs=['Trans_ID','Date','Card_Name','Merchant','Amount','Category','GL_Account','Entity','Reimbursable','Reconciled','Notes']
make_header_row(ws,3,hdrs)

cc_data=[
    ('CC001','05/25/2026','Chase Business','PropWire',299,'Marketing Software',6120,'STLLC','No','No','Monthly sub'),
    ('CC002','05/20/2026','Chase Business','Home Depot',450,'Rehab Materials',5110,'STLLC','No','No','Demo supplies'),
]
for i,row_data in enumerate(cc_data):
    row=4+i
    for j,val in enumerate(row_data):
        c=ws.cell(row=row,column=j+1,value=val)
        c.font=blue_font() if j==4 else data_font()
        if j==4: c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        c.border=thin_border();c.fill=alt_fill() if i%2==1 else white_fill()

set_col_widths(ws,{'A':10,'B':12,'C':18,'D':22,'E':12,'F':22,'G':12,'H':10,'I':12,'J':12,'K':30})
ws.freeze_panes='A4'

# ============================================================
# SHEET 30: CUSTOMERS
# ============================================================
ws=wb.create_sheet("30_CUSTOMERS")
ws.sheet_properties.tabColor=TAB_ARAP
make_title(ws,"CUSTOMER DATABASE",'A1:L1')
ws.row_dimensions[2].height=6

hdrs=['Customer_ID','First_Name','Last_Name','Company','Email','Phone','Address','Customer_Type','Credit_Limit','Outstanding_Balance','Status','Notes']
make_header_row(ws,3,hdrs)

cust_data=[
    ('C001','John','Smith','','john@email.com','314-555-0001','St. Louis, MO','Buyer',500000,0,'Active','Cash buyer — wholesale deals'),
    ('C002','Mary','Johnson','','mary@email.com','314-555-0002','St. Louis, MO','Tenant',2000,0,'Active','Potential tenant — 502 Buckley'),
]
for i,row_data in enumerate(cust_data):
    row=4+i
    for j,val in enumerate(row_data):
        c=ws.cell(row=row,column=j+1,value=val)
        c.font=blue_font() if j in [0,7,10] else data_font()
        if j in [8,9]: c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        c.border=thin_border();c.fill=alt_fill() if i%2==1 else white_fill()

dv=DataValidation(type="list",formula1='"Buyer,Tenant,JV Partner,Investor,Lender"',allow_blank=True)
ws.add_data_validation(dv)
for row in range(4,50): dv.add(ws.cell(row=row,column=8))
set_col_widths(ws,{'A':12,'B':14,'C':14,'D':20,'E':25,'F':16,'G':22,'H':14,'I':14,'J':18,'K':10,'L':35})
ws.freeze_panes='A4'

# ============================================================
# SHEET 31: INVOICES
# ============================================================
ws=wb.create_sheet("31_INVOICES")
ws.sheet_properties.tabColor=TAB_ARAP
make_title(ws,"INVOICE MANAGEMENT",'A1:N1')
ws.row_dimensions[2].height=6

hdrs=['Invoice_ID','Date','Customer_ID','Customer_Name','Property','Description','Amount','Tax','Total','Due_Date','Status','GL_Account','Entity','Notes']
make_header_row(ws,3,hdrs)

inv_data=[
    ('INV-001','05/29/2026','C001','John Smith','','Assignment Fee — 123 Main St',15000,0,'=G4+H4','06/15/2026','Pending',4000,'STLLC',''),
]
for i,row_data in enumerate(inv_data):
    row=4+i
    for j,val in enumerate(row_data):
        c=ws.cell(row=row,column=j+1,value=val)
        c.font=blue_font() if j in [6,7] else data_font()
        if j in [6,7,8]: c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        c.border=thin_border();c.fill=white_fill()

set_col_widths(ws,{'A':12,'B':12,'C':12,'D':22,'E':20,'F':35,'G':14,'H':10,'I':14,'J':12,'K':12,'L':12,'M':10,'N':28})
ws.freeze_panes='A4'

# ============================================================
# SHEET 32: RECEIVABLES
# ============================================================
ws=wb.create_sheet("32_RECEIVABLES")
ws.sheet_properties.tabColor=TAB_ARAP
make_title(ws,"ACCOUNTS RECEIVABLE TRACKER",'A1:K1')
ws.row_dimensions[2].height=6

hdrs=['AR_ID','Invoice_ID','Customer_ID','Customer_Name','Original_Amount','Amount_Paid','Balance_Due','Invoice_Date','Due_Date','Days_Outstanding','Status']
make_header_row(ws,3,hdrs)

ar_data=[('AR001','INV-001','C001','John Smith',15000,0,'=E4-F4','05/29/2026','06/15/2026',f'=IF(K4="Paid",0,MAX(0,TODAY()-I4))','Pending')]
for i,row_data in enumerate(ar_data):
    row=4+i
    for j,val in enumerate(row_data):
        c=ws.cell(row=row,column=j+1,value=val)
        c.font=blue_font() if j in [4,5] else Font(name='Arial',size=10,color=BLACK)
        if j in [4,5,6]: c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        c.border=thin_border();c.fill=white_fill()

set_col_widths(ws,{'A':10,'B':12,'C':12,'D':22,'E':16,'F':14,'G':14,'H':14,'I':12,'J':18,'K':12})
ws.freeze_panes='A4'

# ============================================================
# SHEET 33: AGING_REPORT
# ============================================================
ws=wb.create_sheet("33_AGING_REPORT")
ws.sheet_properties.tabColor=TAB_ARAP
make_title(ws,"ACCOUNTS RECEIVABLE AGING — As of Today",'A1:G1')
ws.row_dimensions[2].height=6

hdrs=['Customer_Name','Current','1-30 Days','31-60 Days','61-90 Days','90+ Days','Total']
make_header_row(ws,3,hdrs)

aging_data=[('John Smith',15000,0,0,0,0)]
for i,row_data in enumerate(aging_data):
    row=4+i
    ws.cell(row=row,column=1,value=row_data[0]).font=data_font()
    for j,val in enumerate(row_data[1:]):
        c=ws.cell(row=row,column=j+2,value=val)
        c.font=blue_font();c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        c.border=thin_border();c.fill=white_fill()
    tc=ws.cell(row=row,column=7,value=f'=SUM(B{row}:F{row})')
    tc.font=Font(name='Arial',size=10,color=BLACK);tc.number_format='$#,##0;($#,##0);"-"';tc.alignment=right_align()
    tc.border=thin_border();tc.fill=white_fill()

total_row=4+len(aging_data)+1
ws.cell(row=total_row,column=1,value='TOTAL').font=bold_font()
for col in range(2,8):
    c=ws.cell(row=total_row,column=col,value=f'=SUM({get_col_letter(col)}4:{get_col_letter(col)}{total_row-1})')
    c.font=bold_font();c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
    c.fill=nav_fill();c.font=Font(name='Arial',size=10,bold=True,color=GOLD);c.border=thin_border()

from openpyxl.utils import get_column_letter
set_col_widths(ws,{'A':28,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14})
ws.freeze_panes='A4'

# ============================================================
# SHEET 40: VENDORS
# ============================================================
ws=wb.create_sheet("40_VENDORS")
ws.sheet_properties.tabColor=TAB_ARAP
make_title(ws,"VENDOR DATABASE",'A1:M1')
ws.row_dimensions[2].height=6

hdrs=['Vendor_ID','Vendor_Name','Contact_Name','Email','Phone','Address','Vendor_Type','Payment_Terms','W9_On_File','1099_Required','Outstanding_Balance','Status','Notes']
make_header_row(ws,3,hdrs)

vendor_data=[
    ('V001','General Contractor','TBD','','314-555-0100','St. Louis, MO','Contractor','Net 30','Yes','Yes',8000,'Active','Primary rehab contractor'),
    ('V002','PropWire','','billing@propwire.com','','','Software','Monthly','No','No',0,'Active','Lead generation software'),
    ('V003','Home Depot','','','','','Supplier','Net 0','No','No',450,'Active','Materials supplier'),
]
for i,row_data in enumerate(vendor_data):
    row=4+i
    for j,val in enumerate(row_data):
        c=ws.cell(row=row,column=j+1,value=val)
        c.font=blue_font() if j in [0,6,11] else data_font()
        if j==10: c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        c.border=thin_border();c.fill=alt_fill() if i%2==1 else white_fill()

dv=DataValidation(type="list",formula1='"Contractor,Supplier,Professional,Lender,Utility,Insurance,Software"',allow_blank=True)
ws.add_data_validation(dv)
for row in range(4,50): dv.add(ws.cell(row=row,column=7))
set_col_widths(ws,{'A':10,'B':24,'C':20,'D':25,'E':16,'F':20,'G':14,'H':14,'I':12,'J':14,'K':18,'L':10,'M':35})
ws.freeze_panes='A4'

# ============================================================
# SHEET 41: BILLS
# ============================================================
ws=wb.create_sheet("41_BILLS")
ws.sheet_properties.tabColor=TAB_ARAP
make_title(ws,"BILL MANAGEMENT — Accounts Payable",'A1:M1')
ws.row_dimensions[2].height=6

hdrs=['Bill_ID','Date','Vendor_ID','Vendor_Name','Description','Amount','Due_Date','Status','GL_Account','Property','Entity','Approved_By','Notes']
make_header_row(ws,3,hdrs)

bill_data=[
    ('BILL-001','05/20/2026','V001','General Contractor','Labor — Demo phase 502 Buckley',8000,'06/20/2026','Unpaid',5100,'502 Buckley','STLLC','Shylow','Phase 1 labor'),
    ('BILL-002','05/25/2026','V002','PropWire','Monthly subscription',299,'06/01/2026','Paid',6120,'','STLLC','Shylow',''),
    ('BILL-003','05/20/2026','V003','Home Depot','Demo supplies',450,'05/20/2026','Paid',5110,'502 Buckley','STLLC','Shylow',''),
]
for i,row_data in enumerate(bill_data):
    row=4+i
    for j,val in enumerate(row_data):
        c=ws.cell(row=row,column=j+1,value=val)
        c.font=blue_font() if j in [5] else data_font()
        if j==5: c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        c.border=thin_border();c.fill=alt_fill() if i%2==1 else white_fill()

set_col_widths(ws,{'A':12,'B':12,'C':10,'D':22,'E':35,'F':12,'G':12,'H':10,'I':12,'J':14,'K':10,'L':14,'M':28})
ws.freeze_panes='A4'

# ============================================================
# SHEET 42: PAYABLES
# ============================================================
ws=wb.create_sheet("42_PAYABLES")
ws.sheet_properties.tabColor=TAB_ARAP
make_title(ws,"ACCOUNTS PAYABLE TRACKER",'A1:K1')
ws.row_dimensions[2].height=6

hdrs=['AP_ID','Bill_ID','Vendor_ID','Vendor_Name','Original_Amount','Amount_Paid','Balance_Due','Bill_Date','Due_Date','Days_Outstanding','Status']
make_header_row(ws,3,hdrs)

ap_data=[
    ('AP001','BILL-001','V001','General Contractor',8000,0,'=E4-F4','05/20/2026','06/20/2026',f'=IF(K4="Paid",0,MAX(0,TODAY()-I4))','Unpaid'),
    ('AP002','BILL-002','V002','PropWire',299,299,'=E5-F5','05/25/2026','06/01/2026',0,'Paid'),
    ('AP003','BILL-003','V003','Home Depot',450,450,'=E6-F6','05/20/2026','05/20/2026',0,'Paid'),
]
for i,row_data in enumerate(ap_data):
    row=4+i
    for j,val in enumerate(row_data):
        c=ws.cell(row=row,column=j+1,value=val)
        c.font=blue_font() if j in [4,5] else Font(name='Arial',size=10,color=BLACK)
        if j in [4,5,6]: c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        c.border=thin_border();c.fill=alt_fill() if i%2==1 else white_fill()

set_col_widths(ws,{'A':10,'B':12,'C':10,'D':22,'E':16,'F':14,'G':14,'H':14,'I':12,'J':18,'K':10})
ws.freeze_panes='A4'

# ============================================================
# SHEET 43: AGING_PAYABLES
# ============================================================
ws=wb.create_sheet("43_AGING_PAYABLES")
ws.sheet_properties.tabColor=TAB_ARAP
make_title(ws,"ACCOUNTS PAYABLE AGING — As of Today",'A1:G1')
ws.row_dimensions[2].height=6

hdrs=['Vendor_Name','Current','1-30 Days','31-60 Days','61-90 Days','90+ Days','Total']
make_header_row(ws,3,hdrs)

ap_aging=[('General Contractor',8000,0,0,0,0),('PropWire',0,0,0,0,0),('Home Depot',0,0,0,0,0)]
for i,row_data in enumerate(ap_aging):
    row=4+i
    ws.cell(row=row,column=1,value=row_data[0]).font=data_font()
    for j,val in enumerate(row_data[1:]):
        c=ws.cell(row=row,column=j+2,value=val)
        c.font=blue_font();c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        c.border=thin_border();c.fill=alt_fill() if i%2==1 else white_fill()
    tc=ws.cell(row=row,column=7,value=f'=SUM(B{row}:F{row})')
    tc.font=Font(name='Arial',size=10,color=BLACK);tc.number_format='$#,##0;($#,##0);"-"'
    tc.alignment=right_align();tc.border=thin_border();tc.fill=alt_fill() if i%2==1 else white_fill()

total_row=4+len(ap_aging)+1
ws.cell(row=total_row,column=1,value='TOTAL').font=bold_font()
ws.cell(row=total_row,column=1).fill=nav_fill()
ws.cell(row=total_row,column=1).font=Font(name='Arial',size=10,bold=True,color=GOLD)
for col in range(2,8):
    from openpyxl.utils import get_column_letter
    c=ws.cell(row=total_row,column=col,value=f'=SUM({get_column_letter(col)}4:{get_column_letter(col)}{total_row-1})')
    c.font=Font(name='Arial',size=10,bold=True,color=GOLD)
    c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
    c.fill=nav_fill();c.border=thin_border()

set_col_widths(ws,{'A':28,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14})
ws.freeze_panes='A4'

wb.save('C:/stllcweb3/Dynasty_Accounting_OS_v1.xlsx')
print("Banking (20-23) and AR/AP (30-43) done")
