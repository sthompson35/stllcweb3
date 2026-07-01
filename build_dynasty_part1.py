"""Dynasty Accounting OS v1.0 - Part 1: Core workbook structure, SYSTEM and ACCOUNTING sheets"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Colors
NAV = "1A2744"   # Navy header bg
GOLD = "D4AF37"  # Gold header text
DARK = "2C3E50"  # Sub-header bg
WHITE = "FFFFFF"
ALT = "F8F9FA"   # Alternating row
BLUE_TEXT = "0000FF"
BLACK = "000000"
GREEN_TEXT = "008000"

# Tab colors
TAB_SYSTEM = "404040"
TAB_ACCT = "1A2744"
TAB_BANK = "008080"
TAB_ARAP = "FF6600"
TAB_RE = "2E7D32"
TAB_DEAL = "6A0DAD"
TAB_INV = "8B0000"
TAB_AI = "0078D4"
TAB_EXEC = "D4AF37"

wb = Workbook()
wb.remove(wb.active)  # remove default sheet

def hdr_font(size=11, bold=True, color=GOLD):
    return Font(name='Arial', size=size, bold=bold, color=color)

def data_font(size=10, bold=False, color=BLACK):
    return Font(name='Arial', size=size, bold=bold, color=color)

def blue_font(size=10):
    return Font(name='Arial', size=10, color=BLUE_TEXT)

def green_font(size=10):
    return Font(name='Arial', size=10, color=GREEN_TEXT)

def nav_fill():
    return PatternFill('solid', fgColor=NAV)

def dark_fill():
    return PatternFill('solid', fgColor=DARK)

def alt_fill():
    return PatternFill('solid', fgColor=ALT)

def center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def make_title(ws, title, span='A1:F1', size=14):
    ws.merge_cells(span)
    c = ws[span.split(':')[0]]
    c.value = title
    c.font = Font(name='Arial', size=size, bold=True, color=GOLD)
    c.fill = nav_fill()
    c.alignment = center()
    ws.row_dimensions[1].height = 30

def make_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start+i, value=h)
        c.font = hdr_font(size=10)
        c.fill = nav_fill()
        c.alignment = center()
        c.border = thin_border()

def set_col_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def style_data_row(ws, row, num_cols, alt=False, col_start=1):
    fill = alt_fill() if alt else PatternFill('solid', fgColor=WHITE)
    for i in range(num_cols):
        c = ws.cell(row=row, column=col_start+i)
        if not c.fill or c.fill.fgColor.rgb == '00000000':
            c.fill = fill
        c.border = thin_border()

# ============================================================
# SHEET 01: SYSTEM_SETTINGS
# ============================================================
ws = wb.create_sheet("01_SYSTEM_SETTINGS")
ws.sheet_properties.tabColor = TAB_SYSTEM

make_title(ws, "DYNASTY ACCOUNTING OS v1.0 — SYSTEM SETTINGS", 'A1:D1')
ws.row_dimensions[2].height = 8

headers = ['Setting', 'Value', 'Notes']
make_header_row(ws, 3, headers)

settings = [
    ('Company Name', 'Shylow Thompson LLC', ''),
    ('Primary Entity', 'Shylow Thompson LLC', ''),
    ('Fiscal Year Start', '01/01/2026', ''),
    ('Fiscal Year End', '12/31/2026', ''),
    ('Accounting Method', 'Accrual', 'GAAP standard'),
    ('Base Currency', 'USD', ''),
    ('Version', '1.0', ''),
    ('Build Date', '05/29/2026', ''),
    ('Next Review', '12/31/2026', ''),
    ('Tax ID', '[ENTER EIN]', 'Federal EIN'),
    ('State', 'Missouri', ''),
    ('Chart of Accounts Version', 'RE-COA-2026', ''),
]

for i, (s, v, n) in enumerate(settings):
    row = 4 + i
    ws.cell(row=row, column=1, value=s).font = data_font()
    vc = ws.cell(row=row, column=2, value=v)
    vc.font = blue_font()
    ws.cell(row=row, column=3, value=n).font = data_font()
    style_data_row(ws, row, 3, alt=(i % 2 == 1))

set_col_widths(ws, {'A': 30, 'B': 30, 'C': 40, 'D': 15})
ws.freeze_panes = 'A4'

# ============================================================
# SHEET 02: USERS_ROLES
# ============================================================
ws = wb.create_sheet("02_USERS_ROLES")
ws.sheet_properties.tabColor = TAB_SYSTEM

make_title(ws, "DYNASTY ACCOUNTING OS — USERS & ROLES", 'A1:H1')
ws.row_dimensions[2].height = 8

hdrs = ['User_ID', 'Full_Name', 'Role', 'Entity_Access', 'Permissions', 'Last_Login', 'Status', 'Notes']
make_header_row(ws, 3, hdrs)

users = [
    ('U001', 'Shylow Thompson', 'CEO / Owner', 'All Entities', 'Full Access', '05/29/2026', 'Active', 'Primary user'),
    ('U002', 'Accountant', 'Controller', 'Shylow Thompson LLC', 'Read/Write', '', 'Active', ''),
    ('U003', 'Investor', 'Investor Portal', 'Assigned Deals Only', 'Read Only', '', 'Active', ''),
]

for i, row_data in enumerate(users):
    row = 4 + i
    for j, val in enumerate(row_data):
        c = ws.cell(row=row, column=j+1, value=val)
        c.font = blue_font() if j in [0,1,2,3,4,6] else data_font()
        c.border = thin_border()
        c.fill = alt_fill() if i % 2 == 1 else PatternFill('solid', fgColor=WHITE)

set_col_widths(ws, {'A': 10, 'B': 22, 'C': 20, 'D': 25, 'E': 18, 'F': 14, 'G': 10, 'H': 25})
ws.freeze_panes = 'A4'

# ============================================================
# SHEET 03: AUDIT_LOG
# ============================================================
ws = wb.create_sheet("03_AUDIT_LOG")
ws.sheet_properties.tabColor = TAB_SYSTEM

make_title(ws, "DYNASTY ACCOUNTING OS — AUDIT LOG", 'A1:J1')
ws.row_dimensions[2].height = 8

hdrs = ['Log_ID', 'Timestamp', 'User_ID', 'User_Name', 'Sheet_Modified', 'Cell_Reference', 'Old_Value', 'New_Value', 'Action_Type', 'Notes']
make_header_row(ws, 3, hdrs)

audit_rows = [
    ('AL001', '05/29/2026 09:00', 'U001', 'Shylow Thompson', '01_SYSTEM_SETTINGS', 'B4', '', 'Shylow Thompson LLC', 'INITIALIZE', 'System initialization'),
    ('AL002', '05/29/2026 09:01', 'U001', 'Shylow Thompson', '50_PROPERTY_MASTER', 'B4', '', '502 Buckley', 'CREATE', 'Added 502 Buckley property'),
]

for i, row_data in enumerate(audit_rows):
    row = 4 + i
    for j, val in enumerate(row_data):
        c = ws.cell(row=row, column=j+1, value=val)
        c.font = data_font()
        c.border = thin_border()
        c.fill = alt_fill() if i % 2 == 1 else PatternFill('solid', fgColor=WHITE)

set_col_widths(ws, {'A': 10, 'B': 18, 'C': 10, 'D': 20, 'E': 22, 'F': 14, 'G': 20, 'H': 20, 'I': 15, 'J': 35})
ws.freeze_panes = 'A4'

# ============================================================
# SHEET 04: ENTITY_MANAGER
# ============================================================
ws = wb.create_sheet("04_ENTITY_MANAGER")
ws.sheet_properties.tabColor = TAB_SYSTEM

make_title(ws, "DYNASTY ACCOUNTING OS — ENTITY MANAGER", 'A1:J1')
ws.row_dimensions[2].height = 8

hdrs = ['Entity_ID', 'Entity_Name', 'Entity_Type', 'EIN', 'State', 'Formation_Date', 'Primary_Contact', 'Role', 'Status', 'Notes']
make_header_row(ws, 3, hdrs)

entities = [
    ('E001', 'Shylow Thompson LLC', 'Operating LLC', '[EIN]', 'Missouri', '', 'Shylow Thompson', 'CEO', 'Active', 'Primary operating entity'),
    ('E002', 'Dynasty OS', 'Tech/SaaS LLC', '[EIN]', 'Missouri', '', 'Shylow Thompson', 'CEO', 'Active', 'Software operations'),
    ('E003', 'KhakiSol', 'Brand/Creative LLC', '[EIN]', 'Missouri', '', 'Shylow Thompson', 'CEO', 'Active', 'Creative & brand'),
    ('E004', '502 Buckley LLC', 'Property LLC', '[EIN]', 'Missouri', '', 'Shylow Thompson', 'Manager', 'Active', 'Property holding entity'),
]

for i, row_data in enumerate(entities):
    row = 4 + i
    for j, val in enumerate(row_data):
        c = ws.cell(row=row, column=j+1, value=val)
        c.font = blue_font() if j in [0,1,2,8] else data_font()
        c.border = thin_border()
        c.fill = alt_fill() if i % 2 == 1 else PatternFill('solid', fgColor=WHITE)

set_col_widths(ws, {'A': 12, 'B': 25, 'C': 20, 'D': 15, 'E': 12, 'F': 16, 'G': 20, 'H': 12, 'I': 10, 'J': 35})
ws.freeze_panes = 'A4'

# ============================================================
# SHEET 05: DASHBOARD_HOME
# ============================================================
ws = wb.create_sheet("05_DASHBOARD_HOME")
ws.sheet_properties.tabColor = TAB_SYSTEM

make_title(ws, "DYNASTY ACCOUNTING OS — COMMAND CENTER", 'A1:H1', size=16)
ws.row_dimensions[1].height = 40

ws.merge_cells('A3:H3')
c = ws['A3']
c.value = "QUICK NAVIGATION"
c.font = hdr_font(size=11)
c.fill = dark_fill()
c.alignment = center()

nav_hdrs = ['Section', 'Sheet Name', 'Description', 'Tab Color']
make_header_row(ws, 4, nav_hdrs)

nav_items = [
    ('SYSTEM', '01_SYSTEM_SETTINGS', 'Global configuration', 'Dark Gray'),
    ('SYSTEM', '02_USERS_ROLES', 'User access management', 'Dark Gray'),
    ('SYSTEM', '03_AUDIT_LOG', 'Change history', 'Dark Gray'),
    ('SYSTEM', '04_ENTITY_MANAGER', 'Business entities', 'Dark Gray'),
    ('SYSTEM', '05_DASHBOARD_HOME', 'This navigation screen', 'Dark Gray'),
    ('ACCOUNTING', '10_CHART_OF_ACCOUNTS', 'Full COA 1000-6999', 'Dark Blue'),
    ('ACCOUNTING', '11_JOURNAL_ENTRIES', 'Double-entry transactions', 'Dark Blue'),
    ('ACCOUNTING', '12_GENERAL_LEDGER', 'Account activity detail', 'Dark Blue'),
    ('ACCOUNTING', '13_TRIAL_BALANCE', 'Debit/credit balance check', 'Dark Blue'),
    ('ACCOUNTING', '14_BALANCE_SHEET', 'Assets = Liabilities + Equity', 'Dark Blue'),
    ('ACCOUNTING', '15_INCOME_STATEMENT', 'Revenue - Expenses = Net Income', 'Dark Blue'),
    ('ACCOUNTING', '16_CASH_FLOW', 'Operating/Investing/Financing', 'Dark Blue'),
    ('ACCOUNTING', '17_BUDGETS', 'Monthly budget by account', 'Dark Blue'),
    ('ACCOUNTING', '18_BUDGET_VS_ACTUAL', 'Variance analysis', 'Dark Blue'),
    ('BANKING', '20_BANK_ACCOUNTS', 'Account registry', 'Teal'),
    ('BANKING', '21_BANK_TRANSACTIONS', 'Transaction feed', 'Teal'),
    ('BANKING', '22_RECONCILIATION', 'Bank rec', 'Teal'),
    ('BANKING', '23_CREDIT_CARDS', 'Card transactions', 'Teal'),
    ('AR/AP', '30_CUSTOMERS', 'Customer database', 'Orange'),
    ('AR/AP', '31_INVOICES', 'Invoice management', 'Orange'),
    ('AR/AP', '32_RECEIVABLES', 'AR tracking', 'Orange'),
    ('AR/AP', '33_AGING_REPORT', 'AR aging buckets', 'Orange'),
    ('AR/AP', '40_VENDORS', 'Vendor database', 'Orange'),
    ('AR/AP', '41_BILLS', 'Bill management', 'Orange'),
    ('AR/AP', '42_PAYABLES', 'AP tracking', 'Orange'),
    ('AR/AP', '43_AGING_PAYABLES', 'AP aging buckets', 'Orange'),
    ('REAL ESTATE', '50_PROPERTY_MASTER', 'Property portfolio', 'Green'),
    ('REAL ESTATE', '51_ACQUISITIONS', 'Deal pipeline', 'Green'),
    ('REAL ESTATE', '52_REHAB_ENGINE', '502 Buckley rehab tracker', 'Green'),
    ('REAL ESTATE', '53_DRAWS', 'Construction draws', 'Green'),
    ('REAL ESTATE', '54_HOLDING_COSTS', 'Monthly holding', 'Green'),
    ('REAL ESTATE', '55_RENTS', 'Rent roll', 'Green'),
    ('REAL ESTATE', '56_CAPEX', 'Capital expenditures', 'Green'),
    ('REAL ESTATE', '57_DISPOSITIONS', 'Property sales', 'Green'),
    ('DEAL ANALYZER', '60_WHOLESALE_ANALYZER', 'Wholesale calculator', 'Purple'),
    ('DEAL ANALYZER', '61_FLIP_ANALYZER', 'Fix & flip calculator', 'Purple'),
    ('DEAL ANALYZER', '62_BRRR_ANALYZER', 'BRRR calculator', 'Purple'),
    ('DEAL ANALYZER', '63_DEVELOPMENT_ANALYZER', 'Ground-up development', 'Purple'),
    ('DEAL ANALYZER', '64_LAND_ANALYZER', 'Land deal calculator', 'Purple'),
    ('INVESTORS', '70_INVESTORS', 'Investor registry', 'Dark Red'),
    ('INVESTORS', '71_CAPITAL_STACK', 'Capital structure', 'Dark Red'),
    ('INVESTORS', '72_WATERFALL', 'Distribution waterfall', 'Dark Red'),
    ('INVESTORS', '73_DISTRIBUTIONS', 'Distribution history', 'Dark Red'),
    ('INVESTORS', '74_INVESTOR_REPORTS', 'Investor summaries', 'Dark Red'),
    ('AI', '80_AI_DEAL_SCORING', 'Deal scoring engine', 'Light Blue'),
    ('AI', '81_AI_FORECASTING', '13-week cash forecast', 'Light Blue'),
    ('AI', '82_AI_ANOMALIES', 'Expense anomaly detection', 'Light Blue'),
    ('AI', '83_AI_ASSISTANT', 'Trooper Accountant AI', 'Light Blue'),
    ('EXECUTIVE', '90_KPI_DASHBOARD', 'CEO cockpit metrics', 'Gold'),
    ('EXECUTIVE', '91_CASH_FORECAST', '12-month cash projection', 'Gold'),
    ('EXECUTIVE', '92_NET_WORTH_TRACKER', 'Net worth history', 'Gold'),
    ('EXECUTIVE', '93_PORTFOLIO_ANALYTICS', 'Portfolio performance', 'Gold'),
    ('EXECUTIVE', '94_EXECUTIVE_REPORT', 'Executive summary report', 'Gold'),
]

for i, (sec, sheet, desc, color) in enumerate(nav_items):
    row = 5 + i
    for j, val in enumerate([sec, sheet, desc, color]):
        c = ws.cell(row=row, column=j+1, value=val)
        c.font = data_font()
        c.border = thin_border()
        c.fill = alt_fill() if i % 2 == 1 else PatternFill('solid', fgColor=WHITE)

status_row = 5 + len(nav_items) + 2
ws.merge_cells(f'A{status_row}:H{status_row}')
c = ws[f'A{status_row}']
c.value = "SYSTEM STATUS"
c.font = hdr_font(size=11)
c.fill = dark_fill()
c.alignment = center()

status_hdrs = ['Metric', 'Value', 'Status']
make_header_row(ws, status_row+1, status_hdrs)

status_items = [
    ('Last Updated', '05/29/2026', 'Current'),
    ('Formula Errors', '0', 'OK'),
    ('Accounting Balance', 'Balanced', 'OK'),
    ('Active Properties', '1', 'Active'),
    ('Active Deals', '1', 'Active'),
    ('Version', '1.0', 'Current'),
]
for i, (m, v, s) in enumerate(status_items):
    row = status_row + 2 + i
    ws.cell(row=row, column=1, value=m).font = data_font()
    ws.cell(row=row, column=2, value=v).font = blue_font()
    ws.cell(row=row, column=3, value=s).font = data_font()
    style_data_row(ws, row, 3, alt=(i % 2 == 1))

set_col_widths(ws, {'A': 18, 'B': 30, 'C': 40, 'D': 15})
ws.freeze_panes = 'A5'

print("SYSTEM sheets (01-05) done")
wb.save('C:/stllcweb3/Dynasty_Accounting_OS_v1.xlsx')
print("Saved part 1")
