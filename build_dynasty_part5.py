"""Dynasty Accounting OS v1.0 - Part 5: DEAL ANALYZERS (60-64), INVESTORS (70-74), AI (80-83), EXECUTIVE (90-94)"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAV="1A2744";GOLD="D4AF37";DARK="2C3E50";WHITE="FFFFFF";ALT="F8F9FA"
BLUE_TEXT="0000FF";BLACK="000000";GREEN_TEXT="008000"
TAB_DEAL="6A0DAD";TAB_INV="8B0000";TAB_AI="0078D4";TAB_EXEC="D4AF37"

wb=load_workbook('C:/stllcweb3/Dynasty_Accounting_OS_v1.xlsx')

def hdr_font(size=10,bold=True,color=GOLD): return Font(name='Arial',size=size,bold=bold,color=color)
def data_font(size=10,color=BLACK): return Font(name='Arial',size=size,color=color)
def blue_font(): return Font(name='Arial',size=10,color=BLUE_TEXT)
def bold_font(size=10,color=BLACK): return Font(name='Arial',size=size,bold=True,color=color)
def green_font(): return Font(name='Arial',size=10,color=GREEN_TEXT)
def nav_fill(): return PatternFill('solid',fgColor=NAV)
def dark_fill(): return PatternFill('solid',fgColor=DARK)
def alt_fill(): return PatternFill('solid',fgColor=ALT)
def white_fill(): return PatternFill('solid',fgColor=WHITE)
def gold_light(): return PatternFill('solid',fgColor='FFF8E1')
def center(): return Alignment(horizontal='center',vertical='center',wrap_text=True)
def left(): return Alignment(horizontal='left',vertical='center',wrap_text=True)
def right_align(): return Alignment(horizontal='right',vertical='center')
def thin_border():
    s=Side(style='thin',color='CCCCCC')
    return Border(left=s,right=s,top=s,bottom=s)
def make_title(ws,title,span='A1:F1',size=13):
    ws.merge_cells(span)
    c=ws[span.split(':')[0]]
    c.value=title;c.font=Font(name='Arial',size=size,bold=True,color=GOLD)
    c.fill=nav_fill();c.alignment=center();ws.row_dimensions[1].height=28
def make_header_row(ws,row,headers,col_start=1):
    for i,h in enumerate(headers):
        c=ws.cell(row=row,column=col_start+i,value=h)
        c.font=hdr_font();c.fill=nav_fill();c.alignment=center();c.border=thin_border()
def set_col_widths(ws,widths):
    for col,w in widths.items(): ws.column_dimensions[col].width=w
def style_row(ws,row,ncols,alt=False,col_start=1):
    fill=alt_fill() if alt else white_fill()
    for i in range(ncols):
        c=ws.cell(row=row,column=col_start+i)
        c.fill=fill;c.border=thin_border()

def analyzer_input(ws,row,label,val,is_pct=False,is_formula=False):
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row,column=1,value=label).font=data_font()
    c=ws.cell(row=row,column=4,value=val)
    if is_formula:
        c.font=Font(name='Arial',size=10,color=BLACK)
    else:
        c.font=blue_font()
    c.number_format='0.0%' if is_pct else '$#,##0;($#,##0);"-"'
    c.alignment=right_align();c.border=thin_border()
    for col in range(1,4): ws.cell(row=row,column=col).border=thin_border()
    style_row(ws,row,4,alt=(row%2==0))
    return row+1

def analyzer_calc(ws,row,label,formula,is_pct=False,is_text=False,is_mult=False):
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row,column=1,value=label).font=bold_font()
    c=ws.cell(row=row,column=4,value=formula)
    c.font=Font(name='Arial',size=10,bold=True,color=BLACK)
    if is_pct: c.number_format='0.0%'
    elif is_mult: c.number_format='0.00x'
    elif not is_text: c.number_format='$#,##0;($#,##0);"-"'
    c.alignment=right_align();c.border=thin_border()
    for col in range(1,4): ws.cell(row=row,column=col).border=thin_border()
    ws.cell(row=row,column=1).fill=dark_fill()
    ws.cell(row=row,column=1).font=Font(name='Arial',size=10,bold=True,color=WHITE)
    for col in range(1,5): ws.cell(row=row,column=col).fill=dark_fill()
    ws.cell(row=row,column=4).fill=dark_fill()
    ws.cell(row=row,column=4).font=Font(name='Arial',size=10,bold=True,color=GOLD)
    return row+1

# ============================================================
# SHEET 60: WHOLESALE_ANALYZER
# ============================================================
ws=wb.create_sheet("60_WHOLESALE_ANALYZER")
ws.sheet_properties.tabColor=TAB_DEAL
make_title(ws,"WHOLESALE DEAL ANALYZER — Dynasty Accounting OS",'A1:F1')
ws.row_dimensions[2].height=6

ws.merge_cells('A3:F3')
c=ws['A3'];c.value="INPUTS (Blue = Editable)";c.font=hdr_font();c.fill=dark_fill();c.alignment=left()

row=4
row=analyzer_input(ws,row,'Property Address','[ENTER ADDRESS]')
row=analyzer_input(ws,row,'ARV (After Repair Value)',150000)
row=analyzer_input(ws,row,'Purchase Price',60000)
row=analyzer_input(ws,row,'Assignment Fee Target',15000)
row=analyzer_input(ws,row,'Marketing Cost',2000)
row=analyzer_input(ws,row,'Inspection Cost',500)
row=analyzer_input(ws,row,'Other Costs',300)

arv_row=5; pp_row=6; fee_row=7; mkt_row=8; ins_row=9; oth_row=10

row+=1
ws.merge_cells(f'A{row}:F{row}')
c=ws.cell(row=row,column=1,value="CALCULATIONS");c.font=hdr_font();c.fill=dark_fill();c.alignment=left()
row+=1

row=analyzer_calc(ws,row,'MAO (70% Rule)',f'=D{arv_row}*0.70',is_text=False)
row=analyzer_calc(ws,row,'Max Purchase Price',f'=D{arv_row}*0.65')
row=analyzer_calc(ws,row,'Gross Spread',f'=D{arv_row}-D{pp_row}')
row=analyzer_calc(ws,row,'Total Costs',f'=D{mkt_row}+D{ins_row}+D{oth_row}')
total_costs_row=row-1
row=analyzer_calc(ws,row,'Net Assignment Fee',f'=D{fee_row}-D{total_costs_row}')
net_fee_row=row-1
row=analyzer_calc(ws,row,'ROI on Marketing',f'=IF(D{total_costs_row}=0,0,D{net_fee_row}/D{total_costs_row})',is_pct=True)
row=analyzer_calc(ws,row,'DECISION',f'=IF(D{pp_row}<=D{arv_row}*0.65,"GO — DEAL WORKS","NO GO — PRICE TOO HIGH")',is_text=True)

row+=2
ws.merge_cells(f'A{row}:F{row}')
c=ws.cell(row=row,column=1,value="WHOLESALE PIPELINE");c.font=hdr_font();c.fill=nav_fill();c.alignment=left()
row+=1

pipe_hdrs=['Lead','Status','ARV','Asking','Offer','Assignment_Fee','Mktg_Cost','Net_Profit','Decision']
make_header_row(ws,row,pipe_hdrs)
row+=1
for i in range(5):
    style_row(ws,row+i,9,alt=(i%2==1))

set_col_widths(ws,{'A':25,'B':20,'C':12,'D':16,'E':14,'F':14,'G':14,'H':14,'I':20})
ws.freeze_panes='A4'

# ============================================================
# SHEET 61: FLIP_ANALYZER
# ============================================================
ws=wb.create_sheet("61_FLIP_ANALYZER")
ws.sheet_properties.tabColor=TAB_DEAL
make_title(ws,"FIX & FLIP ANALYZER — 502 Buckley Analysis",'A1:F1')
ws.row_dimensions[2].height=6

ws.merge_cells('A3:F3')
c=ws['A3'];c.value="INPUTS — 502 Buckley, St. Louis MO";c.font=hdr_font();c.fill=dark_fill();c.alignment=left()

row=4
row=analyzer_input(ws,row,'Property Address','502 Buckley, St. Louis MO 63137')
row=analyzer_input(ws,row,'Purchase Price',114000)
pp_r=row-1
row=analyzer_input(ws,row,'Closing Costs — Purchase',4500)
cc_r=row-1
row=analyzer_input(ws,row,'Rehab Budget',90000)
rh_r=row-1
row=analyzer_input(ws,row,'Holding Period (months)',9)
hp_r=row-1
row=analyzer_input(ws,row,'Monthly Holding Cost',1080)
mh_r=row-1
row=analyzer_input(ws,row,'Selling Price (ARV)',200000)
sp_r=row-1
row=analyzer_input(ws,row,'Realtor Commission %',0.06,is_pct=True)
rc_r=row-1
row=analyzer_input(ws,row,'Closing Costs Sale %',0.03,is_pct=True)
cs_r=row-1
row=analyzer_input(ws,row,'Other Selling Costs',2000)
os_r=row-1

row+=1
ws.merge_cells(f'A{row}:F{row}')
c=ws.cell(row=row,column=1,value="FLIP ANALYSIS");c.font=hdr_font();c.fill=dark_fill();c.alignment=left()
row+=1

row=analyzer_calc(ws,row,'Total Acquisition Cost',f'=D{pp_r}+D{cc_r}')
row=analyzer_calc(ws,row,'Total Rehab Cost',f'=D{rh_r}')
row=analyzer_calc(ws,row,'Total Holding Costs',f'=D{mh_r}*D{hp_r}')
row=analyzer_calc(ws,row,'Total Selling Costs',f'=D{sp_r}*D{rc_r}+D{sp_r}*D{cs_r}+D{os_r}')
total_costs_r=row
row=analyzer_calc(ws,row,'TOTAL PROJECT COST',f'=D{total_costs_r-4}+D{total_costs_r-3}+D{total_costs_r-2}+D{total_costs_r-1}')
tcp_r=row-1
row=analyzer_calc(ws,row,'NET PROFIT',f'=D{sp_r}-D{tcp_r}')
np_r=row-1
row=analyzer_calc(ws,row,'ROI',f'=IF(D{tcp_r}=0,0,D{np_r}/D{tcp_r})',is_pct=True)
roi_r=row-1
row=analyzer_calc(ws,row,'Annualized ROI',f'=IF(D{hp_r}=0,0,(1+D{roi_r})^(12/D{hp_r})-1)',is_pct=True)
row=analyzer_calc(ws,row,'DECISION',f'=IF(D{np_r}>30000,"STRONG FLIP",IF(D{np_r}>15000,"MARGINAL FLIP","DO NOT FLIP"))',is_text=True)

set_col_widths(ws,{'A':28,'B':18,'C':12,'D':16,'E':14,'F':14})
ws.freeze_panes='A4'

# ============================================================
# SHEET 62: BRRR_ANALYZER
# ============================================================
ws=wb.create_sheet("62_BRRR_ANALYZER")
ws.sheet_properties.tabColor=TAB_DEAL
make_title(ws,"BRRR ANALYZER — Buy, Rehab, Rent, Refinance, Repeat",'A1:F1')
ws.row_dimensions[2].height=6

ws.merge_cells('A3:F3')
c=ws['A3'];c.value="INPUTS — 502 Buckley BRRR Analysis";c.font=hdr_font();c.fill=dark_fill();c.alignment=left()

row=4
row=analyzer_input(ws,row,'Purchase Price',114000)
pp_r=row-1
row=analyzer_input(ws,row,'Closing Costs',4500)
cc_r=row-1
row=analyzer_input(ws,row,'Rehab Cost',90000)
rh_r=row-1
row=analyzer_input(ws,row,'ARV After Rehab',200000)
arv_r=row-1
row=analyzer_input(ws,row,'Monthly Rent',1400)
mr_r=row-1
row=analyzer_input(ws,row,'Vacancy Rate %',0.05,is_pct=True)
vac_r=row-1
row=analyzer_input(ws,row,'Operating Expense %',0.35,is_pct=True)
opx_r=row-1
row=analyzer_input(ws,row,'Refinance LTV %',0.75,is_pct=True)
ltv_r=row-1
row=analyzer_input(ws,row,'Refinance Interest Rate %',0.07,is_pct=True)
rate_r=row-1
row=analyzer_input(ws,row,'Refinance Term (years)',30)
term_r=row-1
row=analyzer_input(ws,row,'Total Holding Costs During Rehab',9720)
hc_r=row-1

row+=1
ws.merge_cells(f'A{row}:F{row}')
c=ws.cell(row=row,column=1,value="BRRR ANALYSIS");c.font=hdr_font();c.fill=dark_fill();c.alignment=left()
row+=1

row=analyzer_calc(ws,row,'Total Investment (All-In)',f'=D{pp_r}+D{cc_r}+D{rh_r}+D{hc_r}')
ti_r=row-1
row=analyzer_calc(ws,row,'Refinance Amount',f'=D{arv_r}*D{ltv_r}')
refi_r=row-1
row=analyzer_calc(ws,row,'Monthly Mortgage (Refi)',f'=-PMT(D{rate_r}/12,D{term_r}*12,D{refi_r})')
mm_r=row-1
row=analyzer_calc(ws,row,'Effective Monthly Rent',f'=D{mr_r}*(1-D{vac_r})')
emr_r=row-1
row=analyzer_calc(ws,row,'Monthly Operating Expenses',f'=D{emr_r}*D{opx_r}')
mop_r=row-1
row=analyzer_calc(ws,row,'Monthly NOI',f'=D{emr_r}-D{mop_r}')
noi_r=row-1
row=analyzer_calc(ws,row,'Monthly Cash Flow',f'=D{noi_r}-D{mm_r}')
mcf_r=row-1
row=analyzer_calc(ws,row,'Annual Cash Flow',f'=D{mcf_r}*12')
acf_r=row-1
row=analyzer_calc(ws,row,'Cash Left in Deal',f'=D{ti_r}-D{refi_r}')
cld_r=row-1
row=analyzer_calc(ws,row,'Cash-on-Cash Return',f'=IF(D{cld_r}>0,D{acf_r}/D{cld_r},"INFINITE — Full Refi")',is_text=True)
row=analyzer_calc(ws,row,'Cap Rate',f'=IF(D{arv_r}>0,(D{noi_r}*12)/D{arv_r},0)',is_pct=True)
row=analyzer_calc(ws,row,'DSCR',f'=IF(D{mm_r}>0,D{noi_r}/D{mm_r},0)',is_mult=True)
row=analyzer_calc(ws,row,'Equity Created',f'=D{arv_r}-(D{pp_r}+D{cc_r}+D{rh_r})')
row=analyzer_calc(ws,row,'Cash Recovered',f'=D{refi_r}-(D{pp_r}+D{cc_r}+D{rh_r})')
row=analyzer_calc(ws,row,'DECISION',f'=IF(D{noi_r}/D{mm_r}>=1.25,IF(D{mcf_r}>=0,"STRONG BRRR — GO","BRRR WORKS — WEAK CASH FLOW"),"DSCR TOO LOW — RENEGOTIATE")',is_text=True)

set_col_widths(ws,{'A':32,'B':18,'C':12,'D':18,'E':14,'F':14})
ws.freeze_panes='A4'

# ============================================================
# SHEET 63: DEVELOPMENT_ANALYZER
# ============================================================
ws=wb.create_sheet("63_DEVELOPMENT_ANALYZER")
ws.sheet_properties.tabColor=TAB_DEAL
make_title(ws,"DEVELOPMENT DEAL ANALYZER",'A1:F1')

ws.merge_cells('A3:F3')
c=ws['A3'];c.value="INPUTS";c.font=hdr_font();c.fill=dark_fill();c.alignment=left()

row=4
row=analyzer_input(ws,row,'Land Cost',0)
land_r=row-1
row=analyzer_input(ws,row,'Soft Costs (permits, arch, eng)',0)
soft_r=row-1
row=analyzer_input(ws,row,'Hard Construction Costs',0)
hard_r=row-1
row=analyzer_input(ws,row,'Financing Costs',0)
fin_r=row-1
row=analyzer_input(ws,row,'Contingency %',0.10,is_pct=True)
cont_r=row-1
row=analyzer_input(ws,row,'Number of Units',1)
units_r=row-1
row=analyzer_input(ws,row,'Projected Sale Price per Unit',0)
sale_r=row-1

row+=1
ws.merge_cells(f'A{row}:F{row}')
c=ws.cell(row=row,column=1,value="DEVELOPMENT ANALYSIS");c.font=hdr_font();c.fill=dark_fill();c.alignment=left()
row+=1

row=analyzer_calc(ws,row,'Base Development Cost',f'=D{land_r}+D{soft_r}+D{hard_r}+D{fin_r}')
base_r=row-1
row=analyzer_calc(ws,row,'Contingency Reserve',f'=D{base_r}*D{cont_r}')
cont_amt_r=row-1
row=analyzer_calc(ws,row,'TOTAL DEVELOPMENT COST',f'=D{base_r}+D{cont_amt_r}')
tdc_r=row-1
row=analyzer_calc(ws,row,'Cost Per Unit',f'=IF(D{units_r}>0,D{tdc_r}/D{units_r},0)')
row=analyzer_calc(ws,row,'Gross Development Value (GDV)',f'=D{sale_r}*D{units_r}')
gdv_r=row-1
row=analyzer_calc(ws,row,'Development Profit',f'=D{gdv_r}-D{tdc_r}')
dp_r=row-1
row=analyzer_calc(ws,row,'Development Margin',f'=IF(D{gdv_r}=0,0,D{dp_r}/D{gdv_r})',is_pct=True)
dm_r=row-1
row=analyzer_calc(ws,row,'ROI',f'=IF(D{tdc_r}=0,0,D{dp_r}/D{tdc_r})',is_pct=True)
row=analyzer_calc(ws,row,'DECISION',f'=IF(D{dm_r}>0.2,"VIABLE PROJECT",IF(D{dm_r}>0.1,"MARGINAL","NOT VIABLE"))',is_text=True)

set_col_widths(ws,{'A':30,'B':18,'C':12,'D':18,'E':14,'F':14})
ws.freeze_panes='A4'

# ============================================================
# SHEET 64: LAND_ANALYZER
# ============================================================
ws=wb.create_sheet("64_LAND_ANALYZER")
ws.sheet_properties.tabColor=TAB_DEAL
make_title(ws,"LAND DEAL ANALYZER",'A1:F1')

ws.merge_cells('A3:F3')
c=ws['A3'];c.value="INPUTS";c.font=hdr_font();c.fill=dark_fill();c.alignment=left()

row=4
row=analyzer_input(ws,row,'Purchase Price',0)
pp_r=row-1
row=analyzer_input(ws,row,'Entitlement Costs',0)
ent_r=row-1
row=analyzer_input(ws,row,'Carrying Costs / Month',0)
carry_r=row-1
row=analyzer_input(ws,row,'Hold Period (months)',12)
hold_r=row-1
row=analyzer_input(ws,row,'Target Sale Price',0)
sp_r=row-1
row=analyzer_input(ws,row,'Assignment Fee Option',0)
af_r=row-1

row+=1
ws.merge_cells(f'A{row}:F{row}')
c=ws.cell(row=row,column=1,value="LAND ANALYSIS");c.font=hdr_font();c.fill=dark_fill();c.alignment=left()
row+=1

row=analyzer_calc(ws,row,'Total Carrying Costs',f'=D{carry_r}*D{hold_r}')
tc_r=row-1
row=analyzer_calc(ws,row,'Total Cost (Hold)',f'=D{pp_r}+D{ent_r}+D{tc_r}')
tch_r=row-1
row=analyzer_calc(ws,row,'Net Profit (Sell)',f'=D{sp_r}-D{tch_r}')
np_r=row-1
row=analyzer_calc(ws,row,'ROI (Sell)',f'=IF(D{tch_r}=0,0,D{np_r}/D{tch_r})',is_pct=True)
row=analyzer_calc(ws,row,'Net Profit (Assign)',f'=D{af_r}-(D{ent_r}+D{carry_r}*3)')
na_r=row-1
row=analyzer_calc(ws,row,'ROI (Assign)',f'=IF(D{pp_r}=0,0,D{na_r}/D{pp_r})',is_pct=True)
row=analyzer_calc(ws,row,'Annualized ROI',f'=IF(D{hold_r}=0,0,(1+D{np_r}/IF(D{tch_r}=0,1,D{tch_r}))^(12/D{hold_r})-1)',is_pct=True)

set_col_widths(ws,{'A':30,'B':18,'C':12,'D':18,'E':14,'F':14})
ws.freeze_panes='A4'

# ============================================================
# SHEET 70: INVESTORS
# ============================================================
ws=wb.create_sheet("70_INVESTORS")
ws.sheet_properties.tabColor=TAB_INV
make_title(ws,"INVESTOR REGISTRY — Dynasty Accounting OS",'A1:N1')
ws.row_dimensions[2].height=6

hdrs=['Investor_ID','Full_Name','Entity','Email','Phone','Accredited','Capital_Committed','Capital_Called','Capital_Returned','Pref_Return_%','Profit_Split_%','Current_Balance','Status','Notes']
make_header_row(ws,3,hdrs)

inv_data=[
    ('INV001','Private Lender 1','[TBD]','','','Yes',50000,0,0,0.08,0.30,0,'Active','Hard money / private lender potential'),
]
for i,row_data in enumerate(inv_data):
    row=4+i
    for j,val in enumerate(row_data):
        c=ws.cell(row=row,column=j+1,value=val)
        c.font=blue_font() if j in [6,7,8] else data_font()
        if j in [6,7,8,11]: c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        elif j in [9,10]: c.number_format='0.0%';c.alignment=right_align()
        c.border=thin_border();c.fill=alt_fill() if i%2==1 else white_fill()

set_col_widths(ws,{'A':12,'B':22,'C':18,'D':25,'E':16,'F':12,'G':18,'H':16,'I':18,'J':14,'K':14,'L':16,'M':10,'N':35})
ws.freeze_panes='A4'

# ============================================================
# SHEET 71: CAPITAL_STACK
# ============================================================
ws=wb.create_sheet("71_CAPITAL_STACK")
ws.sheet_properties.tabColor=TAB_INV
make_title(ws,"CAPITAL STACK — 502 Buckley",'A1:K1')
ws.row_dimensions[2].height=6

hdrs=['Tranche','Source','Type','Amount','Rate_%','Term','LTV_%','Position','Monthly_Payment','Total_Cost','Notes']
make_header_row(ws,3,hdrs)

cs_data=[
    ('Senior Debt','FHA Mortgage','Debt',114000,0.065,'30 years','=D4/200000','1st Lien',750,'=I4*360','FHA purchase loan at 6.5%'),
    ('Equity','Owner Capital','Equity',18500,0,'N/A','N/A','Equity',0,'N/A','Owner equity + closing costs'),
]
for i,row_data in enumerate(cs_data):
    row=4+i
    for j,val in enumerate(row_data):
        c=ws.cell(row=row,column=j+1,value=val)
        if j in [3,8]:
            c.font=blue_font() if isinstance(val,(int,float)) else Font(name='Arial',size=10,color=BLACK)
            c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        elif j==4: c.font=blue_font();c.number_format='0.0%';c.alignment=right_align()
        elif j==6: c.font=Font(name='Arial',size=10,color=BLACK);c.number_format='0.0%';c.alignment=right_align()
        else: c.font=data_font()
        c.border=thin_border();c.fill=alt_fill() if i%2==1 else white_fill()

# Fix LTV formula display
ws['G4'].value='=D4/200000'
ws['G4'].number_format='0.0%';ws['G4'].alignment=right_align()
ws['G4'].font=Font(name='Arial',size=10,color=BLACK)
ws['J4'].value='=I4*360'
ws['J4'].number_format='$#,##0;($#,##0);"-"';ws['J4'].alignment=right_align()
ws['J4'].font=Font(name='Arial',size=10,color=BLACK)

total_row=4+len(cs_data)+1
ws.cell(row=total_row,column=1,value='TOTALS').font=bold_font()
ws.cell(row=total_row,column=1).fill=nav_fill()
ws.cell(row=total_row,column=1).font=Font(name='Arial',size=10,bold=True,color=GOLD)
for col,formula in [(4,f'=SUM(D4:D{total_row-1})'),(9,f'=SUM(I4:I{total_row-1})')]:
    c=ws.cell(row=total_row,column=col,value=formula)
    c.font=Font(name='Arial',size=10,bold=True,color=GOLD);c.fill=nav_fill()
    c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align();c.border=thin_border()
for col in range(1,12):
    if ws.cell(row=total_row,column=col).fill.fgColor.rgb in ('00000000',):
        ws.cell(row=total_row,column=col).fill=nav_fill()
    ws.cell(row=total_row,column=col).border=thin_border()

set_col_widths(ws,{'A':14,'B':18,'C':10,'D':16,'E':10,'F':10,'G':10,'H':10,'I':16,'J':16,'K':35})
ws.freeze_panes='A4'

# ============================================================
# SHEET 72: WATERFALL
# ============================================================
ws=wb.create_sheet("72_WATERFALL")
ws.sheet_properties.tabColor=TAB_INV
make_title(ws,"DISTRIBUTION WATERFALL — 3-Tier Structure",'A1:D1')

ws.merge_cells('A3:D3')
c=ws['A3'];c.value="WATERFALL INPUTS";c.font=hdr_font();c.fill=dark_fill();c.alignment=left()

inputs=[
    (4,'Total Distributions Available',0),
    (5,'Total Investor Capital Invested',50000),
    (6,'Preferred Return Rate %',0.08),
    (7,'Investor Profit Share %',0.30),
    (8,'Owner Profit Share %',0.70),
]
for row,label,val in inputs:
    ws.cell(row=row,column=1,value=label).font=data_font()
    c=ws.cell(row=row,column=4,value=val)
    c.font=blue_font()
    c.number_format='0.0%' if isinstance(val,float) and val<1 else '$#,##0;($#,##0);"-"'
    c.alignment=right_align();c.border=thin_border()
    for col in range(1,4): ws.cell(row=row,column=col).border=thin_border()
    style_row(ws,row,4,alt=(row%2==0))

ws.row_dimensions[9].height=10
ws.merge_cells('A10:D10')
c=ws['A10'];c.value="WATERFALL DISTRIBUTION";c.font=hdr_font();c.fill=nav_fill();c.alignment=left()

wf_steps=[
    (11,'TIER 1: Return of Capital','=MIN(D4,D5)',False),
    (12,'Remaining After Tier 1','=MAX(0,D4-D11)',False),
    (13,'TIER 2: Preferred Return (8% annualized)','=MIN(D12,D5*D6)',False),
    (14,'Remaining After Tier 2','=MAX(0,D12-D13)',False),
    (15,'TIER 3: Investor Profit Share (30%)','=D14*D7',False),
    (16,'Owner Profit Share (70%)','=D14*D8',False),
    (17,'','',''),
    (18,'INVESTOR TOTAL','=D11+D13+D15',False),
    (19,'OWNER TOTAL','=D16',False),
    (20,'GRAND TOTAL CHECK','=D18+D19',False),
]
for row,label,formula,is_pct in wf_steps:
    if label=='': continue
    is_tier=label.startswith('TIER') or label.endswith('TOTAL') or label.endswith('CHECK')
    ws.cell(row=row,column=1,value=label).font=bold_font() if is_tier else data_font()
    if is_tier:
        ws.merge_cells(f'A{row}:C{row}')
        ws.cell(row=row,column=1).fill=dark_fill()
        ws.cell(row=row,column=1).font=Font(name='Arial',size=10,bold=True,color=GOLD)
    else:
        ws.merge_cells(f'A{row}:C{row}')
    c=ws.cell(row=row,column=4,value=formula)
    c.font=Font(name='Arial',size=10,bold=is_tier,color=GOLD if is_tier else BLACK)
    c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
    if is_tier: c.fill=dark_fill()
    for col in range(1,5): ws.cell(row=row,column=col).border=thin_border()
    if not is_tier: style_row(ws,row,4,alt=(row%2==0))

set_col_widths(ws,{'A':30,'B':16,'C':16,'D':18})

# ============================================================
# SHEET 73: DISTRIBUTIONS
# ============================================================
ws=wb.create_sheet("73_DISTRIBUTIONS")
ws.sheet_properties.tabColor=TAB_INV
make_title(ws,"DISTRIBUTION HISTORY",'A1:L1')
ws.row_dimensions[2].height=6

hdrs=['Dist_ID','Date','Property_Fund','Investor_ID','Investor_Name','Dist_Type','Amount','ROC','Preferred','Profit','Cumulative','Notes']
make_header_row(ws,3,hdrs)
ws.cell(row=4,column=1,value='[No distributions yet]').font=Font(name='Arial',size=10,italic=True,color='888888')
ws.merge_cells('A4:L4')

set_col_widths(ws,{'A':10,'B':12,'C':18,'D':12,'E':22,'F':22,'G':14,'H':14,'I':12,'J':12,'K':14,'L':28})
ws.freeze_panes='A4'

# ============================================================
# SHEET 74: INVESTOR_REPORTS
# ============================================================
ws=wb.create_sheet("74_INVESTOR_REPORTS")
ws.sheet_properties.tabColor=TAB_INV
make_title(ws,"INVESTOR SUMMARY REPORT — Dynasty Accounting OS",'A1:H1')

ws.merge_cells('A3:H3')
c=ws['A3'];c.value=f"Report Date: 05/29/2026  |  Prepared by: Dynasty Accounting OS v1.0";c.font=data_font()
c.fill=alt_fill();c.alignment=left()

ws.row_dimensions[4].height=8
hdrs=['Investor_ID','Name','Capital_Invested','Distributions','Current_Balance','Pref_Return_Earned','IRR','Status']
make_header_row(ws,5,hdrs)

ws.cell(row=6,column=1,value='INV001').font=data_font()
ws.cell(row=6,column=2,value='Private Lender 1').font=data_font()
for j,val in enumerate([50000,0,50000,0,'N/A','Active']):
    c=ws.cell(row=6,column=3+j,value=val)
    c.font=blue_font() if j in [0,1,2,3] else data_font()
    if j<4: c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
    c.border=thin_border();c.fill=white_fill()
style_row(ws,6,8)

set_col_widths(ws,{'A':12,'B':22,'C':18,'D':16,'E':18,'F':20,'G':10,'H':12})
ws.freeze_panes='A6'

# ============================================================
# SHEET 80: AI_DEAL_SCORING
# ============================================================
ws=wb.create_sheet("80_AI_DEAL_SCORING")
ws.sheet_properties.tabColor=TAB_AI
make_title(ws,"AI DEAL SCORING ENGINE — TROOPER_ACCOUNTANT",'A1:M1')

ws.merge_cells('A3:M3')
c=ws['A3']
c.value="STATUS: Placeholder formulas active | Connect AI API for enhanced ML scoring | All scores calculated via weighted formula matrix"
c.font=Font(name='Arial',size=10,italic=True,color='888888');c.fill=alt_fill();c.alignment=left()

hdrs=['Deal_ID','Address','Deal_Type','ARV','Purchase','Rehab','Spread','Margin_%','DSCR','Cash_Flow','Score','Grade','AI_Recommendation']
make_header_row(ws,5,hdrs)

ws.cell(row=6,column=1,value='DEAL001').font=data_font()
ws.cell(row=6,column=2,value='502 Buckley, St. Louis MO').font=data_font()
ws.cell(row=6,column=3,value='BRRR').font=data_font()
for j,val in enumerate([200000,114000,90000]):
    c=ws.cell(row=6,column=4+j,value=val)
    c.font=blue_font();c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
    c.border=thin_border();c.fill=white_fill()
ws.cell(row=6,column=7,value='=D6-E6').font=Font(name='Arial',size=10,color=BLACK)
ws.cell(row=6,column=7).number_format='$#,##0;($#,##0);"-"';ws.cell(row=6,column=7).alignment=right_align()
ws.cell(row=6,column=8,value='=IF(D6>0,(D6-E6-F6)/D6,0)').font=Font(name='Arial',size=10,color=BLACK)
ws.cell(row=6,column=8).number_format='0.0%';ws.cell(row=6,column=8).alignment=right_align()
ws.cell(row=6,column=9,value=0).font=blue_font()
ws.cell(row=6,column=10,value=0).font=blue_font()
ws.cell(row=6,column=11,value='=MAX(0,MIN(100,H6*100*0.4+IF(I6>=1.25,30,I6*24)+IF(J6>0,30,0)))').font=Font(name='Arial',size=10,color=BLACK)
ws.cell(row=6,column=11).number_format='0'
ws.cell(row=6,column=12,value='=IF(K6>=80,"A",IF(K6>=70,"B",IF(K6>=60,"C","D")))').font=Font(name='Arial',size=10,color=BLACK)
ws.cell(row=6,column=13,value='=IF(L6="A","STRONG BUY",IF(L6="B","REVIEW",IF(L6="C","PASS","AVOID")))').font=Font(name='Arial',size=10,color=BLACK)
style_row(ws,6,13)

set_col_widths(ws,{'A':10,'B':28,'C':14,'D':14,'E':14,'F':14,'G':14,'H':12,'I':10,'J':12,'K':8,'L':8,'M':20})
ws.freeze_panes='A5'

# ============================================================
# SHEET 81: AI_FORECASTING
# ============================================================
ws=wb.create_sheet("81_AI_FORECASTING")
ws.sheet_properties.tabColor=TAB_AI
make_title(ws,"13-WEEK ROLLING CASH FORECAST — Dynasty Accounting OS",'A1:I1')
ws.row_dimensions[2].height=6

hdrs=['Week','Week_Start','Week_End','Beginning_Balance','Expected_Inflows','Expected_Outflows','Net_Change','Ending_Balance','Alert']
make_header_row(ws,3,hdrs)

# 13 weeks of forecast data
import datetime
start=datetime.date(2026,5,29)
for i in range(13):
    row=4+i
    wk_start=start+datetime.timedelta(weeks=i)
    wk_end=wk_start+datetime.timedelta(days=6)
    ws.cell(row=row,column=1,value=i+1).font=data_font();ws.cell(row=row,column=1).alignment=center()
    ws.cell(row=row,column=2,value=wk_start.strftime('%m/%d/%Y')).font=data_font()
    ws.cell(row=row,column=3,value=wk_end.strftime('%m/%d/%Y')).font=data_font()
    if i==0:
        beg=ws.cell(row=row,column=4,value=17200)
        beg.font=blue_font()
    else:
        beg=ws.cell(row=row,column=4,value=f'=H{row-1}')
        beg.font=green_font()
    beg.number_format='$#,##0;($#,##0);"-"';beg.alignment=right_align()
    inflow=ws.cell(row=row,column=5,value=0 if i<2 else 5000)
    inflow.font=blue_font();inflow.number_format='$#,##0;($#,##0);"-"';inflow.alignment=right_align()
    outflow_vals={0:3800,1:1500,2:8500,3:1500,4:8500,5:1500,6:8500,7:1500,8:8500,9:1500,10:3800,11:1500,12:1500}
    out=ws.cell(row=row,column=6,value=outflow_vals.get(i,1500))
    out.font=blue_font();out.number_format='$#,##0;($#,##0);"-"';out.alignment=right_align()
    nc=ws.cell(row=row,column=7,value=f'=E{row}-F{row}')
    nc.font=Font(name='Arial',size=10,color=BLACK);nc.number_format='$#,##0;($#,##0);"-"';nc.alignment=right_align()
    end=ws.cell(row=row,column=8,value=f'=D{row}+G{row}')
    end.font=Font(name='Arial',size=10,color=BLACK);end.number_format='$#,##0;($#,##0);"-"';end.alignment=right_align()
    alert=ws.cell(row=row,column=9,value=f'=IF(H{row}<5000,"CRITICAL — LOW CASH",IF(H{row}<15000,"WARNING — MONITOR","OK"))')
    alert.font=Font(name='Arial',size=10,color=BLACK)
    style_row(ws,row,9,alt=(i%2==1))

set_col_widths(ws,{'A':8,'B':14,'C':14,'D':18,'E':18,'F':18,'G':14,'H':18,'I':25})
ws.freeze_panes='A4'

# ============================================================
# SHEET 82: AI_ANOMALIES
# ============================================================
ws=wb.create_sheet("82_AI_ANOMALIES")
ws.sheet_properties.tabColor=TAB_AI
make_title(ws,"AI ANOMALY DETECTION — Expense Monitor",'A1:I1')

ws.merge_cells('A3:I3')
c=ws['A3'];c.value="STATUS: Monitoring active | Connect to live transaction feed for real-time anomaly detection | Threshold: >50% variance from average"
c.font=Font(name='Arial',size=10,italic=True,color='888888');c.fill=alt_fill();c.alignment=left()

hdrs=['Detection_Date','Account','Category','Amount','Avg_Amount','Variance_%','Anomaly_Flag','Severity','Notes']
make_header_row(ws,5,hdrs)

anom_data=[
    ('05/25/2026',6120,'Marketing Software',299,299,0),
    ('05/20/2026',5110,'Rehab Materials',3500,3500,0),
]
for i,row_data in enumerate(anom_data):
    row=6+i
    date,acct,cat,amt,avg,var=row_data
    ws.cell(row=row,column=1,value=date).font=data_font()
    ws.cell(row=row,column=2,value=acct).font=data_font()
    ws.cell(row=row,column=3,value=cat).font=data_font()
    ac=ws.cell(row=row,column=4,value=amt);ac.font=blue_font();ac.number_format='$#,##0;($#,##0);"-"';ac.alignment=right_align()
    ag=ws.cell(row=row,column=5,value=avg);ag.font=blue_font();ag.number_format='$#,##0;($#,##0);"-"';ag.alignment=right_align()
    vp=ws.cell(row=row,column=6,value=f'=IF(E{row}=0,0,(D{row}-E{row})/E{row})')
    vp.font=Font(name='Arial',size=10,color=BLACK);vp.number_format='0.0%';vp.alignment=right_align()
    af=ws.cell(row=row,column=7,value=f'=IF(ABS(F{row})>0.5,"FLAGGED","Normal")')
    af.font=Font(name='Arial',size=10,color=BLACK)
    sv=ws.cell(row=row,column=8,value=f'=IF(ABS(F{row})>1.0,"HIGH",IF(ABS(F{row})>0.5,"MEDIUM","LOW"))')
    sv.font=Font(name='Arial',size=10,color=BLACK)
    style_row(ws,row,9,alt=(i%2==1))

set_col_widths(ws,{'A':14,'B':12,'C':22,'D':14,'E':14,'F':14,'G':16,'H':10,'I':35})
ws.freeze_panes='A5'

# ============================================================
# SHEET 83: AI_ASSISTANT
# ============================================================
ws=wb.create_sheet("83_AI_ASSISTANT")
ws.sheet_properties.tabColor=TAB_AI
make_title(ws,"AI ACCOUNTING ASSISTANT — TROOPER_ACCOUNTANT",'A1:F1',size=12)

ws.merge_cells('A3:F3')
c=ws['A3'];c.value="TROOPER_ACCOUNTANT | Institutional-grade AI accounting assistant for Dynasty Accounting OS"
c.font=hdr_font(size=10);c.fill=dark_fill();c.alignment=left()

capabilities=[
    ('Auto-categorize transactions','Connect to bank feed API — auto-matches transactions to COA'),
    ('Detect expense anomalies','Flags variances >50% from historical average per category'),
    ('Forecast 13-week cash flow','Uses budget + AR/AP aging to project weekly cash position'),
    ('Analyze property performance','Calculates ROI, DSCR, Cap Rate, Cash-on-Cash for all properties'),
    ('Score deals (0-100)','Weighted scoring: margin 40%, DSCR 30%, cash flow 30%'),
    ('Prepare lender packages','Generates formatted lender-ready financial summaries'),
    ('Prepare investor reports','Generates distribution statements and portfolio reports'),
    ('Monitor KPI drift','Alerts when KPIs deviate >10% from targets'),
    ('Generate financial narratives','Creates plain-English summaries of financial performance'),
    ('Simulate scenarios','Models best/base/worst case scenarios for each property'),
]

ws.merge_cells('A5:C5')
ws.cell(row=5,column=1,value='CAPABILITY').font=hdr_font();ws.cell(row=5,column=1).fill=nav_fill()
ws.cell(row=5,column=1).alignment=center()
ws.merge_cells('D5:F5')
ws.cell(row=5,column=4,value='DESCRIPTION').font=hdr_font();ws.cell(row=5,column=4).fill=nav_fill()
ws.cell(row=5,column=4).alignment=center()
for col in range(1,7): ws.cell(row=5,column=col).border=thin_border()

for i,(cap,desc) in enumerate(capabilities):
    row=6+i
    ws.merge_cells(f'A{row}:C{row}')
    ws.cell(row=row,column=1,value=cap).font=bold_font()
    ws.cell(row=row,column=1).alignment=left()
    ws.merge_cells(f'D{row}:F{row}')
    ws.cell(row=row,column=4,value=desc).font=data_font()
    ws.cell(row=row,column=4).alignment=left()
    style_row(ws,row,6,alt=(i%2==1))

api_row=6+len(capabilities)+2
ws.merge_cells(f'A{api_row}:F{api_row}')
c=ws.cell(row=api_row,column=1,value="API INTEGRATION CONFIGURATION");c.font=hdr_font();c.fill=nav_fill();c.alignment=left()

api_hdrs=['Function','Status','API_Endpoint','Last_Run','Notes']
make_header_row(ws,api_row+1,api_hdrs)

api_fns=['Transaction Categorization','Anomaly Detection','Cash Forecasting','Deal Scoring','Report Generation']
for i,fn in enumerate(api_fns):
    row=api_row+2+i
    ws.cell(row=row,column=1,value=fn).font=data_font()
    ws.cell(row=row,column=2,value='PLACEHOLDER — Configure API endpoint').font=Font(name='Arial',size=10,color='FF6600')
    for col in range(1,6): ws.cell(row=row,column=col).border=thin_border()
    style_row(ws,row,5,alt=(i%2==1))

set_col_widths(ws,{'A':22,'B':22,'C':22,'D':22,'E':22,'F':22})
ws.freeze_panes='A6'

# ============================================================
# SHEET 90: KPI_DASHBOARD
# ============================================================
ws=wb.create_sheet("90_KPI_DASHBOARD")
ws.sheet_properties.tabColor=TAB_EXEC
make_title(ws,"DYNASTY ACCOUNTING OS — KPI COMMAND CENTER",'A1:H1',size=14)
ws.row_dimensions[1].height=38

def kpi_box(ws,row,col,label,value,note='',is_formula=False,fmt='currency'):
    ws.merge_cells(f'{chr(64+col)}{row}:{chr(64+col+1)}{row}')
    cell=ws.cell(row=row,column=col,value=label)
    cell.font=Font(name='Arial',size=9,bold=True,color='888888');cell.alignment=center()
    cell.fill=PatternFill('solid',fgColor=DARK);cell.border=thin_border()
    ws.merge_cells(f'{chr(64+col)}{row+1}:{chr(64+col+1)}{row+1}')
    vcell=ws.cell(row=row+1,column=col,value=value)
    vcell.font=Font(name='Arial',size=16,bold=True,color=BLACK)
    vcell.fill=PatternFill('solid',fgColor=WHITE);vcell.alignment=center()
    vcell.border=thin_border()
    if fmt=='currency': vcell.number_format='$#,##0;($#,##0);"-"'
    elif fmt=='pct': vcell.number_format='0.0%'
    elif fmt=='mult': vcell.number_format='0.00x'
    if not is_formula: vcell.font=Font(name='Arial',size=16,bold=True,color=BLUE_TEXT)
    if note:
        ws.merge_cells(f'{chr(64+col)}{row+2}:{chr(64+col+1)}{row+2}')
        nc=ws.cell(row=row+2,column=col,value=note)
        nc.font=Font(name='Arial',size=8,italic=True,color='888888');nc.alignment=center()
        nc.fill=PatternFill('solid',fgColor=ALT);nc.border=thin_border()
    ws.row_dimensions[row].height=16;ws.row_dimensions[row+1].height=32;ws.row_dimensions[row+2].height=14

row_start=3
ws.merge_cells(f'A{row_start}:H{row_start}')
c=ws[f'A{row_start}'];c.value="CASH & FINANCIAL METRICS";c.font=hdr_font();c.fill=dark_fill();c.alignment=left()
row_start+=1

kpi_box(ws,row_start,1,'CASH AVAILABLE',17200,'Bank balance',False)
kpi_box(ws,row_start,3,'MONTHLY REVENUE',0,'MTD revenue',False)
kpi_box(ws,row_start,5,'MONTHLY EXPENSES',11799,'MTD expenses',False)
kpi_box(ws,row_start,7,'NET INCOME MTD',-11799,'Revenue - Expenses',False)

row_start+=3
ws.merge_cells(f'A{row_start}:H{row_start}')
c=ws[f'A{row_start}'];c.value="REAL ESTATE PORTFOLIO METRICS";c.font=hdr_font();c.fill=dark_fill();c.alignment=left()
row_start+=1

kpi_box(ws,row_start,1,'ACTIVE PROPERTIES',1,'Count',False,'count')
kpi_box(ws,row_start,3,'PORTFOLIO ARV',200000,'Total ARV',False)
kpi_box(ws,row_start,5,'PORTFOLIO EQUITY',86000,'ARV - Debt',False)
kpi_box(ws,row_start,7,'REHAB IN PROGRESS',1,'Properties',False,'count')

row_start+=3
ws.merge_cells(f'A{row_start}:H{row_start}')
c=ws[f'A{row_start}'];c.value="OBLIGATIONS & PIPELINE";c.font=hdr_font();c.fill=dark_fill();c.alignment=left()
row_start+=1

kpi_box(ws,row_start,1,'TOTAL DEBT',114000,'Mortgages + loans',False)
kpi_box(ws,row_start,3,'MONTHLY DEBT SERVICE',750,'All loan payments',False)
kpi_box(ws,row_start,5,'AP BALANCE',8000,'Total owed to vendors',False)
kpi_box(ws,row_start,7,'AR BALANCE',15000,'Total owed to company',False)

row_start+=3
ws.merge_cells(f'A{row_start}:H{row_start}')
c=ws[f'A{row_start}'];c.value="DEAL PIPELINE";c.font=hdr_font();c.fill=dark_fill();c.alignment=left()
row_start+=1

kpi_box(ws,row_start,1,'WHOLESALE PIPELINE','$0','Active deals',False,'text')
kpi_box(ws,row_start,3,'FLIP PIPELINE','$114K','502 Buckley',False,'text')
kpi_box(ws,row_start,5,'BRRR PIPELINE','$114K','502 Buckley',False,'text')
kpi_box(ws,row_start,7,'MTD ASSIGN FEES','$0','This month',False,'currency')

set_col_widths(ws,{'A':14,'B':14,'C':14,'D':14,'E':14,'F':14,'G':14,'H':14})

# ============================================================
# SHEET 91: CASH_FORECAST
# ============================================================
ws=wb.create_sheet("91_CASH_FORECAST")
ws.sheet_properties.tabColor=TAB_EXEC
make_title(ws,"12-MONTH CASH FORECAST — FY 2026",'A1:G1')
ws.row_dimensions[2].height=6

hdrs=['Month','Beginning_Balance','Revenue','Expenses','Net_Change','Ending_Balance','Notes']
make_header_row(ws,3,hdrs)

months=['Jan 2026','Feb 2026','Mar 2026','Apr 2026','May 2026','Jun 2026','Jul 2026','Aug 2026','Sep 2026','Oct 2026','Nov 2026','Dec 2026']
rev_vals=[0,0,0,0,500,5000,5000,5000,5000,10000,10000,10000]
exp_vals=[0,0,0,0,11799,11799,11799,5000,5000,5000,5000,5000]

for i,(month,rev,exp) in enumerate(zip(months,rev_vals,exp_vals)):
    row=4+i
    ws.cell(row=row,column=1,value=month).font=data_font()
    if i==0:
        beg=ws.cell(row=row,column=2,value=0)
        beg.font=blue_font()
    else:
        beg=ws.cell(row=row,column=2,value=f'=F{row-1}')
        beg.font=green_font()
    beg.number_format='$#,##0;($#,##0);"-"';beg.alignment=right_align()
    rc=ws.cell(row=row,column=3,value=rev);rc.font=blue_font()
    rc.number_format='$#,##0;($#,##0);"-"';rc.alignment=right_align()
    ec=ws.cell(row=row,column=4,value=exp);ec.font=blue_font()
    ec.number_format='$#,##0;($#,##0);"-"';ec.alignment=right_align()
    nc=ws.cell(row=row,column=5,value=f'=C{row}-D{row}')
    nc.font=Font(name='Arial',size=10,color=BLACK);nc.number_format='$#,##0;($#,##0);"-"';nc.alignment=right_align()
    end=ws.cell(row=row,column=6,value=f'=B{row}+E{row}')
    end.font=Font(name='Arial',size=10,color=BLACK);end.number_format='$#,##0;($#,##0);"-"';end.alignment=right_align()
    notes='Rehab in progress' if i<5 else ('Est. wholesale activity' if i<9 else 'Projected')
    ws.cell(row=row,column=7,value=notes).font=data_font()
    style_row(ws,row,7,alt=(i%2==1))

set_col_widths(ws,{'A':14,'B':18,'C':14,'D':14,'E':14,'F':18,'G':28})
ws.freeze_panes='A4'

# ============================================================
# SHEET 92: NET_WORTH_TRACKER
# ============================================================
ws=wb.create_sheet("92_NET_WORTH_TRACKER")
ws.sheet_properties.tabColor=TAB_EXEC
make_title(ws,"NET WORTH TRACKER — Shylow Thompson LLC",'A1:I1')
ws.row_dimensions[2].height=6

hdrs=['Date','Cash_Total','Property_Values','Other_Assets','Total_Assets','Total_Liabilities','Net_Worth','Change','Notes']
make_header_row(ws,3,hdrs)

row=4
ws.cell(row=row,column=1,value='05/29/2026').font=blue_font()
ws.cell(row=row,column=2,value=17200).font=blue_font()
ws.cell(row=row,column=2).number_format='$#,##0;($#,##0);"-"';ws.cell(row=row,column=2).alignment=right_align()
ws.cell(row=row,column=3,value=200000).font=blue_font()
ws.cell(row=row,column=3).number_format='$#,##0;($#,##0);"-"';ws.cell(row=row,column=3).alignment=right_align()
ws.cell(row=row,column=4,value=0).font=blue_font()
ws.cell(row=row,column=4).number_format='$#,##0;($#,##0);"-"';ws.cell(row=row,column=4).alignment=right_align()
ws.cell(row=row,column=5,value='=B4+C4+D4').font=Font(name='Arial',size=10,color=BLACK)
ws.cell(row=row,column=5).number_format='$#,##0;($#,##0);"-"';ws.cell(row=row,column=5).alignment=right_align()
ws.cell(row=row,column=6,value=122299).font=blue_font()
ws.cell(row=row,column=6).number_format='$#,##0;($#,##0);"-"';ws.cell(row=row,column=6).alignment=right_align()
ws.cell(row=row,column=7,value='=E4-F4').font=Font(name='Arial',size=10,color=BLACK)
ws.cell(row=row,column=7).number_format='$#,##0;($#,##0);"-"';ws.cell(row=row,column=7).alignment=right_align()
ws.cell(row=row,column=8,value=0).font=blue_font()
ws.cell(row=row,column=8).number_format='$#,##0;($#,##0);"-"';ws.cell(row=row,column=8).alignment=right_align()
ws.cell(row=row,column=9,value='Inception — Dynasty Accounting OS v1.0').font=data_font()
style_row(ws,row,9)

set_col_widths(ws,{'A':14,'B':16,'C':18,'D':14,'E':16,'H':14,'F':18,'G':16,'I':35})
ws.freeze_panes='A4'

# ============================================================
# SHEET 93: PORTFOLIO_ANALYTICS
# ============================================================
ws=wb.create_sheet("93_PORTFOLIO_ANALYTICS")
ws.sheet_properties.tabColor=TAB_EXEC
make_title(ws,"PORTFOLIO ANALYTICS — Dynasty Accounting OS",'A1:J1')

ws.merge_cells('A3:J3')
c=ws['A3'];c.value="PROPERTY PERFORMANCE TABLE";c.font=hdr_font();c.fill=dark_fill();c.alignment=left()
hdrs=['Property','Purchase','ARV','Equity','NOI','DSCR','Cap_Rate','Cash_Flow','Status','Grade']
make_header_row(ws,4,hdrs)

row=5
props=[('502 Buckley',114000,200000,'=C5-B5-114000',0,0,0,0,'Active — Rehab','B')]
for prop_data in props:
    for j,val in enumerate(prop_data):
        c=ws.cell(row=row,column=j+1,value=val)
        if j in [1,2]: c.font=blue_font();c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        elif j==3: c.font=Font(name='Arial',size=10,color=BLACK);c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        elif j in [4,7]: c.font=blue_font();c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
        elif j==5: c.font=blue_font();c.number_format='0.00x';c.alignment=right_align()
        elif j==6: c.font=blue_font();c.number_format='0.0%';c.alignment=right_align()
        else: c.font=data_font()
        c.border=thin_border();c.fill=white_fill()
    row+=1

row+=1
ws.merge_cells(f'A{row}:J{row}')
c=ws.cell(row=row,column=1,value="REVENUE MIX (YTD)");c.font=hdr_font();c.fill=dark_fill();c.alignment=left()
row+=1
rev_mix=[('Wholesale Assignment Fees',0),('Flip Profits',0),('Rental Income',0),('Consulting/Other',500)]
for name,val in rev_mix:
    ws.cell(row=row,column=1,value=name).font=data_font()
    c=ws.cell(row=row,column=2,value=val);c.font=blue_font();c.number_format='$#,##0;($#,##0);"-"';c.alignment=right_align()
    for col in range(1,3): ws.cell(row=row,column=col).border=thin_border()
    style_row(ws,row,2,alt=(row%2==0))
    row+=1

total_rev_row=row
ws.cell(row=row,column=1,value='TOTAL REVENUE YTD').font=bold_font()
ws.cell(row=row,column=2,value=f'=SUM(B{total_rev_row-4}:B{total_rev_row-1})').font=bold_font()
ws.cell(row=row,column=2).number_format='$#,##0;($#,##0);"-"';ws.cell(row=row,column=2).alignment=right_align()
row+=2

ws.merge_cells(f'A{row}:J{row}')
c=ws.cell(row=row,column=1,value="PORTFOLIO KPIs");c.font=hdr_font();c.fill=nav_fill();c.alignment=left()
row+=1
kpis=[('Total Properties',1),('Total Portfolio ARV','$200,000'),('Total Debt','$114,000'),('Total Equity','$86,000'),('Avg Cap Rate','0.0%'),('Portfolio DSCR','0.00x')]
for kname,kval in kpis:
    ws.cell(row=row,column=1,value=kname).font=data_font()
    ws.cell(row=row,column=2,value=kval).font=blue_font()
    for col in range(1,3): ws.cell(row=row,column=col).border=thin_border()
    style_row(ws,row,2,alt=(row%2==0))
    row+=1

set_col_widths(ws,{'A':28,'B':18,'C':16,'D':14,'E':14,'F':12,'G':12,'H':14,'I':18,'J':12})
ws.freeze_panes='A4'

# ============================================================
# SHEET 94: EXECUTIVE_REPORT
# ============================================================
ws=wb.create_sheet("94_EXECUTIVE_REPORT")
ws.sheet_properties.tabColor=TAB_EXEC
make_title(ws,"DYNASTY ACCOUNTING OS — EXECUTIVE SUMMARY REPORT",'A1:F1',size=13)

ws.merge_cells('A2:F2')
c=ws['A2'];c.value="Shylow Thompson LLC  |  Period: YTD 2026  |  Generated: 05/29/2026  |  Dynasty Accounting OS v1.0"
c.font=Font(name='Arial',size=10,italic=True,color='888888');c.fill=alt_fill();c.alignment=center()

sections=[
    (4,"1. COMPANY OVERVIEW",[
        ("Entity","Shylow Thompson LLC — Real Estate Investment Company"),
        ("State","Missouri"),
        ("Business Lines","Wholesale, BRRR, Flip, Development, Rentals"),
        ("Accounting System","Dynasty Accounting OS v1.0 — Accrual Basis"),
        ("Reporting Period","January 1, 2026 — May 29, 2026"),
    ]),
    (11,"2. FINANCIAL HIGHLIGHTS",[
        ("Total Revenue YTD","$500"),
        ("Total Expenses YTD","$11,799"),
        ("Net Income YTD","($11,299)"),
        ("Cash Position","$17,200"),
        ("Note","Net loss reflects startup + rehab investment phase — expected"),
    ]),
    (18,"3. BALANCE SHEET SUMMARY",[
        ("Total Assets","$217,200"),
        ("Total Liabilities","$122,299"),
        ("Net Worth / Equity","$94,901"),
        ("Primary Asset","502 Buckley — ARV $200,000"),
    ]),
    (24,"4. REAL ESTATE PORTFOLIO",[
        ("Active Properties","1 — 502 Buckley, St. Louis MO 63137"),
        ("Status","Active Rehabilitation"),
        ("Purchase Price","$114,000"),
        ("Rehab Budget","$90,000"),
        ("Rehab Actual","$11,500 (12.8% complete)"),
        ("Estimated ARV","$200,000"),
        ("Exit Strategy","BRRR — Refinance + Hold as Rental"),
    ]),
    (33,"5. KEY RISKS & OPPORTUNITIES",[
        ("Risk 1","Rehab cost overruns — monitor weekly vs 90K budget"),
        ("Risk 2","Cash position — 17K available vs 90K rehab remaining"),
        ("Opportunity 1","502 Buckley BRRR — est. $86K equity creation"),
        ("Opportunity 2","Wholesale pipeline — 0 active deals, needs marketing"),
    ]),
    (40,"6. 90-DAY OUTLOOK",[
        ("Rehab Completion Target","August 2026 (est.)"),
        ("Estimated Cash Need","$78,500 additional rehab funding"),
        ("Revenue Opportunity","$10,000-$25,000 wholesale deals"),
        ("Next Action","Secure rehab financing / private lender capital"),
    ]),
]

for start_row,title,items in sections:
    ws.merge_cells(f'A{start_row}:F{start_row}')
    c=ws.cell(row=start_row,column=1,value=title)
    c.font=hdr_font(size=11);c.fill=dark_fill();c.alignment=left();c.border=thin_border()
    for i,(label,val) in enumerate(items):
        row=start_row+1+i
        ws.cell(row=row,column=1,value='   '+label).font=bold_font()
        ws.merge_cells(f'B{row}:F{row}')
        ws.cell(row=row,column=2,value=val).font=data_font()
        ws.cell(row=row,column=2).alignment=left()
        style_row(ws,row,6,alt=(i%2==1))
        for col in range(1,7): ws.cell(row=row,column=col).border=thin_border()

disc_row=48
ws.merge_cells(f'A{disc_row}:F{disc_row}')
ws.cell(row=disc_row,column=1,value='DISCLAIMER: This report is for internal management use only. Generated by Dynasty Accounting OS v1.0 | Shylow Thompson LLC Confidential')
ws.cell(row=disc_row,column=1).font=Font(name='Arial',size=8,italic=True,color='888888')
ws.cell(row=disc_row,column=1).alignment=center()
ws.cell(row=disc_row,column=1).fill=alt_fill()

set_col_widths(ws,{'A':22,'B':20,'C':18,'D':16,'E':14,'F':14})

wb.save('C:/stllcweb3/Dynasty_Accounting_OS_v1.xlsx')
print("Deal Analyzers, Investors, AI, and Executive sheets done")
print("All 40 sheets created!")
