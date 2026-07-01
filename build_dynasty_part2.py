"""Dynasty Accounting OS v1.0 - Part 2: ACCOUNTING CORE sheets 10-18"""
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

NAV = "1A2744"; GOLD = "D4AF37"; DARK = "2C3E50"; WHITE = "FFFFFF"; ALT = "F8F9FA"
BLUE_TEXT = "0000FF"; BLACK = "000000"; GREEN_TEXT = "008000"
TAB_ACCT = "1A2744"

wb = load_workbook('C:/stllcweb3/Dynasty_Accounting_OS_v1.xlsx')

def hdr_font(size=10, bold=True, color=GOLD): return Font(name='Arial', size=size, bold=bold, color=color)
def data_font(size=10, bold=False, color=BLACK): return Font(name='Arial', size=size, bold=bold, color=color)
def blue_font(): return Font(name='Arial', size=10, color=BLUE_TEXT)
def green_font(): return Font(name='Arial', size=10, color=GREEN_TEXT)
def bold_font(size=10, color=BLACK): return Font(name='Arial', size=size, bold=True, color=color)
def nav_fill(): return PatternFill('solid', fgColor=NAV)
def dark_fill(): return PatternFill('solid', fgColor=DARK)
def alt_fill(): return PatternFill('solid', fgColor=ALT)
def white_fill(): return PatternFill('solid', fgColor=WHITE)
def gold_fill(): return PatternFill('solid', fgColor="FFF8E1")
def center(): return Alignment(horizontal='center', vertical='center', wrap_text=True)
def left(): return Alignment(horizontal='left', vertical='center', wrap_text=True)
def right_align(): return Alignment(horizontal='right', vertical='center')
def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)
def make_title(ws, title, span='A1:F1', size=13):
    ws.merge_cells(span)
    c = ws[span.split(':')[0]]
    c.value = title; c.font = Font(name='Arial', size=size, bold=True, color=GOLD)
    c.fill = nav_fill(); c.alignment = center(); ws.row_dimensions[1].height = 28
def make_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start+i, value=h)
        c.font = hdr_font(); c.fill = nav_fill(); c.alignment = center(); c.border = thin_border()
def set_col_widths(ws, widths):
    for col, w in widths.items(): ws.column_dimensions[col].width = w
def style_row(ws, row, ncols, alt=False, col_start=1):
    fill = alt_fill() if alt else white_fill()
    for i in range(ncols):
        c = ws.cell(row=row, column=col_start+i)
        if not c.fill or c.fill.fgColor.rgb in ('00000000','FFFFFFFF','00FFFFFF'):
            c.fill = fill
        c.border = thin_border()

# ============================================================
# SHEET 10: CHART_OF_ACCOUNTS
# ============================================================
ws = wb.create_sheet("10_CHART_OF_ACCOUNTS")
ws.sheet_properties.tabColor = TAB_ACCT
make_title(ws, "CHART OF ACCOUNTS — Dynasty Accounting OS v1.0", 'A1:H1')
ws.row_dimensions[2].height = 6

hdrs = ['Account_Number', 'Account_Name', 'Account_Type', 'Sub_Type', 'Normal_Balance', 'Description', 'Active', 'Entity']
make_header_row(ws, 3, hdrs)

coa = [
    # ASSETS
    (1000,'Cash — Operating Account','Asset','Current Asset','Debit','Primary operating checking','Yes','All'),
    (1010,'Cash — Savings Account','Asset','Current Asset','Debit','Business savings','Yes','All'),
    (1020,'Cash — Reserve Fund','Asset','Current Asset','Debit','Capital reserves','Yes','All'),
    (1030,'Cash — Tax Reserve','Asset','Current Asset','Debit','Tax payment reserves','Yes','All'),
    (1040,'Petty Cash','Asset','Current Asset','Debit','Small cash purchases','Yes','All'),
    (1100,'Accounts Receivable','Asset','Current Asset','Debit','Money owed to company','Yes','All'),
    (1110,'AR — Assignment Fees','Asset','Current Asset','Debit','Wholesale assignment receivables','Yes','All'),
    (1120,'AR — Rental Income','Asset','Current Asset','Debit','Rent receivables','Yes','All'),
    (1200,'Prepaid Expenses','Asset','Current Asset','Debit','Prepaid insurance, rent','Yes','All'),
    (1210,'Prepaid Insurance','Asset','Current Asset','Debit','','Yes','All'),
    (1300,'Inventory — Land','Asset','Real Estate Inventory','Debit','Land held for sale','Yes','All'),
    (1310,'Inventory — Wholesale Contracts','Asset','Real Estate Inventory','Debit','Contracts under assignment','Yes','All'),
    (1400,'Real Estate — Investment Properties','Asset','Fixed Asset','Debit','Properties held for investment','Yes','All'),
    (1410,'502 Buckley — Building','Asset','Fixed Asset','Debit','Building value at purchase','Yes','STLLC'),
    (1420,'502 Buckley — Land','Asset','Fixed Asset','Debit','Land value at purchase','Yes','STLLC'),
    (1430,'Accumulated Depreciation — Buildings','Asset','Fixed Asset','Credit','Contra asset','Yes','All'),
    (1500,'Equipment','Asset','Fixed Asset','Debit','Tools, equipment','Yes','All'),
    (1510,'Vehicles','Asset','Fixed Asset','Debit','Company vehicles','Yes','All'),
    (1520,'Accumulated Depreciation — Equipment','Asset','Fixed Asset','Credit','Contra asset','Yes','All'),
    (1600,'Security Deposits Held','Asset','Other Asset','Debit','Tenant deposits','Yes','All'),
    (1700,'Due from Related Entities','Asset','Other Asset','Debit','Inter-company receivables','Yes','All'),
    (1800,'Investments — Joint Ventures','Asset','Other Asset','Debit','JV investment interests','Yes','All'),
    (1900,'Other Assets','Asset','Other Asset','Debit','Miscellaneous','Yes','All'),
    # LIABILITIES
    (2000,'Accounts Payable','Liability','Current Liability','Credit','Vendor bills owed','Yes','All'),
    (2010,'AP — Contractors','Liability','Current Liability','Credit','Contractor invoices','Yes','All'),
    (2020,'AP — Materials','Liability','Current Liability','Credit','Material supplier invoices','Yes','All'),
    (2100,'Credit Cards Payable','Liability','Current Liability','Credit','Credit card balances','Yes','All'),
    (2110,'Chase Business Card','Liability','Current Liability','Credit','','Yes','All'),
    (2200,'Accrued Expenses','Liability','Current Liability','Credit','Expenses incurred not yet paid','Yes','All'),
    (2210,'Accrued Interest','Liability','Current Liability','Credit','','Yes','All'),
    (2220,'Accrued Payroll','Liability','Current Liability','Credit','','Yes','All'),
    (2300,'Security Deposits — Tenant','Liability','Current Liability','Credit','Deposits held for tenants','Yes','All'),
    (2400,'Short-Term Debt','Liability','Current Liability','Credit','Debt due within 12 months','Yes','All'),
    (2500,'Mortgage Payable — 502 Buckley','Liability','Long-Term Liability','Credit','FHA mortgage principal','Yes','STLLC'),
    (2510,'Hard Money Loans','Liability','Long-Term Liability','Credit','Hard money borrowings','Yes','All'),
    (2520,'Private Lender Notes','Liability','Long-Term Liability','Credit','Private money loans','Yes','All'),
    (2530,'Seller Financing','Liability','Long-Term Liability','Credit','Owner carry notes','Yes','All'),
    (2600,'Deferred Revenue','Liability','Long-Term Liability','Credit','Advance rent payments','Yes','All'),
    (2700,'Due to Related Entities','Liability','Long-Term Liability','Credit','Inter-company payables','Yes','All'),
    (2800,'Tax Payable — Income','Liability','Current Liability','Credit','Estimated income taxes','Yes','All'),
    (2810,'Tax Payable — Sales','Liability','Current Liability','Credit','Sales tax collected','Yes','All'),
    (2900,'Other Liabilities','Liability','Other Liability','Credit','Miscellaneous','Yes','All'),
    # EQUITY
    (3000,'Members Equity — Shylow Thompson','Equity','Owners Equity','Credit','Primary member equity','Yes','All'),
    (3010,'Retained Earnings','Equity','Retained Earnings','Credit','Accumulated profits','Yes','All'),
    (3020,'Members Draws','Equity','Owners Equity','Debit','Owner distributions','Yes','All'),
    (3030,'Capital Contributions','Equity','Owners Equity','Credit','Cash invested by owner','Yes','All'),
    (3100,'Investor Capital — Preferred','Equity','Investor Equity','Credit','Preferred equity investors','Yes','All'),
    (3110,'Investor Capital — Common','Equity','Investor Equity','Credit','Common equity investors','Yes','All'),
    (3200,'Profit Distributions','Equity','Distributions','Debit','Investor profit distributions','Yes','All'),
    # REVENUE
    (4000,'Wholesale Assignment Fees','Revenue','Operating Revenue','Credit','Assignment contract fees','Yes','All'),
    (4010,'Flip Profits','Revenue','Operating Revenue','Credit','Net profit from flips','Yes','All'),
    (4020,'Rental Income — Residential','Revenue','Operating Revenue','Credit','Residential rents','Yes','All'),
    (4030,'Rental Income — Commercial','Revenue','Operating Revenue','Credit','Commercial rents','Yes','All'),
    (4040,'Late Fees','Revenue','Operating Revenue','Credit','Tenant late fees','Yes','All'),
    (4050,'Application Fees','Revenue','Operating Revenue','Credit','Rental application fees','Yes','All'),
    (4100,'Consulting Income','Revenue','Other Revenue','Credit','Advisory/consulting fees','Yes','All'),
    (4110,'Dynasty OS Revenue','Revenue','Other Revenue','Credit','SaaS/software revenue','Yes','Dynasty OS'),
    (4120,'KhakiSol Revenue','Revenue','Other Revenue','Credit','Creative/brand revenue','Yes','KhakiSol'),
    (4200,'Interest Income','Revenue','Other Revenue','Credit','Bank and investment interest','Yes','All'),
    (4300,'Gain on Sale — Real Estate','Revenue','Other Revenue','Credit','Profit from property sales','Yes','All'),
    (4400,'Other Income','Revenue','Other Revenue','Credit','Miscellaneous income','Yes','All'),
    # COGS
    (5000,'Cost of Properties Sold','COGS','Direct Cost','Debit','Book value of sold properties','Yes','All'),
    (5010,'Acquisition Costs','COGS','Direct Cost','Debit','Purchase price + closing costs','Yes','All'),
    (5100,'Rehab — Labor','COGS','Direct Cost','Debit','Contractor labor','Yes','All'),
    (5110,'Rehab — Materials','COGS','Direct Cost','Debit','Materials and supplies','Yes','All'),
    (5120,'Rehab — Permits','COGS','Direct Cost','Debit','Building permits','Yes','All'),
    (5130,'Rehab — Subcontractors','COGS','Direct Cost','Debit','Specialty contractors','Yes','All'),
    (5200,'Holding Costs','COGS','Direct Cost','Debit','While under rehab/sale','Yes','All'),
    (5210,'Holding — Mortgage Interest','COGS','Direct Cost','Debit','Interest during hold','Yes','All'),
    (5220,'Holding — Insurance','COGS','Direct Cost','Debit','Property insurance','Yes','All'),
    (5230,'Holding — Utilities','COGS','Direct Cost','Debit','Utilities during hold','Yes','All'),
    (5240,'Holding — Property Tax','COGS','Direct Cost','Debit','Taxes during hold','Yes','All'),
    (5300,'Assignment Costs','COGS','Direct Cost','Debit','Marketing, finding fees','Yes','All'),
    (5400,'Development Costs','COGS','Direct Cost','Debit','Ground-up development','Yes','All'),
    # EXPENSES
    (6000,'Salaries & Wages','Expense','Operating Expense','Debit','Payroll','Yes','All'),
    (6010,'Contract Labor','Expense','Operating Expense','Debit','1099 workers','Yes','All'),
    (6100,'Marketing — Digital','Expense','Operating Expense','Debit','Online ads, PPC','Yes','All'),
    (6110,'Marketing — Direct Mail','Expense','Operating Expense','Debit','Mailers, postcards','Yes','All'),
    (6120,'Marketing — Software (PropWire/DealMachine)','Expense','Operating Expense','Debit','Lead gen tools','Yes','All'),
    (6130,'Marketing — Cold Calling','Expense','Operating Expense','Debit','Dialers, call centers','Yes','All'),
    (6200,'Office Expenses','Expense','Operating Expense','Debit','Supplies, equipment','Yes','All'),
    (6210,'Rent / Lease','Expense','Operating Expense','Debit','Office space','Yes','All'),
    (6220,'Utilities — Office','Expense','Operating Expense','Debit','Electric, internet','Yes','All'),
    (6300,'Professional Fees — Legal','Expense','Operating Expense','Debit','Attorney fees','Yes','All'),
    (6310,'Professional Fees — Accounting','Expense','Operating Expense','Debit','CPA/bookkeeper','Yes','All'),
    (6320,'Professional Fees — Consulting','Expense','Operating Expense','Debit','Advisors','Yes','All'),
    (6400,'Software & Technology','Expense','Operating Expense','Debit','Subscriptions','Yes','All'),
    (6410,'Dynasty OS Development','Expense','Operating Expense','Debit','App/web development','Yes','Dynasty OS'),
    (6500,'Insurance','Expense','Operating Expense','Debit','Business insurance','Yes','All'),
    (6510,'E&O Insurance','Expense','Operating Expense','Debit','Errors & Omissions','Yes','All'),
    (6600,'Vehicle & Travel','Expense','Operating Expense','Debit','Car, mileage, travel','Yes','All'),
    (6610,'Fuel','Expense','Operating Expense','Debit','Gas expenses','Yes','All'),
    (6700,'Bank Fees','Expense','Operating Expense','Debit','Account fees, wire fees','Yes','All'),
    (6800,'Interest Expense','Expense','Operating Expense','Debit','Loan interest','Yes','All'),
    (6810,'Mortgage Interest — 502 Buckley','Expense','Operating Expense','Debit','FHA mortgage interest','Yes','STLLC'),
    (6900,'Depreciation Expense','Expense','Operating Expense','Debit','Annual depreciation','Yes','All'),
    (6910,'Depreciation — 502 Buckley','Expense','Operating Expense','Debit','','Yes','STLLC'),
    (6990,'Miscellaneous Expenses','Expense','Operating Expense','Debit','Other expenses','Yes','All'),
]

section_colors = {'Asset': 'E3F2FD', 'Liability': 'FCE4EC', 'Equity': 'F3E5F5',
                  'Revenue': 'E8F5E9', 'COGS': 'FFF8E1', 'Expense': 'FBE9E7'}

for i, row_data in enumerate(coa):
    row = 4 + i
    sec_color = section_colors.get(str(row_data[2]), ALT)
    fill = PatternFill('solid', fgColor=sec_color)
    for j, val in enumerate(row_data):
        c = ws.cell(row=row, column=j+1, value=val)
        c.font = data_font()
        c.fill = fill
        c.border = thin_border()
        if j == 0:
            c.font = bold_font()
            c.alignment = center()

set_col_widths(ws, {'A': 16, 'B': 42, 'C': 12, 'D': 22, 'E': 14, 'F': 38, 'G': 8, 'H': 12})
ws.freeze_panes = 'A4'

# ============================================================
# SHEET 11: JOURNAL_ENTRIES
# ============================================================
ws = wb.create_sheet("11_JOURNAL_ENTRIES")
ws.sheet_properties.tabColor = TAB_ACCT
make_title(ws, "JOURNAL ENTRIES — Double-Entry Accounting Ledger", 'A1:M1')
ws.row_dimensions[2].height = 6

hdrs = ['JE_ID', 'Date', 'Description', 'Debit_Acct', 'Debit_Name', 'Debit_Amt', 'Credit_Acct', 'Credit_Name', 'Credit_Amt', 'Entity', 'Reference', 'Posted', 'Notes']
make_header_row(ws, 3, hdrs)

je_data = [
    ('JE001', '05/01/2026', 'Opening Capital Contribution', 1000, 'Cash — Operating Account', 25000, 3030, 'Capital Contributions', 25000, 'STLLC', 'OPENING', 'Yes', 'Initial capital'),
    ('JE002', '05/01/2026', 'Purchase 502 Buckley — Building', 1410, '502 Buckley — Building', 97900, 2500, 'Mortgage Payable — 502 Buckley', 114000, 'STLLC', 'CLOSING-502', 'Yes', 'FHA purchase close'),
    ('JE003', '05/01/2026', 'Purchase 502 Buckley — Land', 1420, '502 Buckley — Land', 16100, 1000, 'Cash — Operating Account', 4500, 'STLLC', 'CLOSING-502', 'Yes', 'Land allocation + closing costs'),
    ('JE004', '05/01/2026', 'Acquisition Closing Costs', 5010, 'Acquisition Costs', 4500, 1000, 'Cash — Operating Account', 0, 'STLLC', 'CLOSING-502', 'Yes', 'Included in JE003 credit'),
    ('JE005', '05/15/2026', 'Rehab Materials — Demo Phase', 5110, 'Rehab — Materials', 3500, 1000, 'Cash — Operating Account', 3500, 'STLLC', 'INV-001', 'Yes', 'Demo phase materials'),
    ('JE006', '05/20/2026', 'Rehab Labor — Contractor Draw 1', 5100, 'Rehab — Labor', 8000, 2010, 'AP — Contractors', 8000, 'STLLC', 'INV-002', 'Yes', 'First labor draw'),
    ('JE007', '05/25/2026', 'PropWire Subscription', 6120, 'Marketing — Software', 299, 2110, 'Chase Business Card', 299, 'STLLC', 'CC-001', 'Yes', 'Monthly subscription'),
]

for i, row_data in enumerate(je_data):
    row = 4 + i
    for j, val in enumerate(row_data):
        c = ws.cell(row=row, column=j+1, value=val)
        if j in [5, 8]:
            c.font = blue_font()
            c.number_format = '$#,##0;($#,##0);"-"'
            c.alignment = right_align()
        elif j == 11:
            c.font = blue_font()
        else:
            c.font = data_font()
        c.border = thin_border()
        c.fill = alt_fill() if i % 2 == 1 else white_fill()

# Balance check row
check_row = 4 + len(je_data) + 2
ws.cell(row=check_row, column=1, value='BALANCE CHECK (Posted):').font = bold_font()
ws.cell(row=check_row, column=5, value='Total Debits:').font = bold_font()
ws.cell(row=check_row, column=6, value='=SUMIF(L4:L100,"Yes",F4:F100)').font = Font(name='Arial', size=10, bold=True, color=BLACK)
ws.cell(row=check_row, column=6).number_format = '$#,##0;($#,##0);"-"'
ws.cell(row=check_row, column=8, value='Total Credits:').font = bold_font()
ws.cell(row=check_row, column=9, value='=SUMIF(L4:L100,"Yes",I4:I100)').font = Font(name='Arial', size=10, bold=True, color=BLACK)
ws.cell(row=check_row, column=9).number_format = '$#,##0;($#,##0);"-"'
ws.cell(row=check_row+1, column=5, value='Difference (must = 0):').font = bold_font()
diff = ws.cell(row=check_row+1, column=6, value=f'=F{check_row}-I{check_row}')
diff.font = bold_font()
diff.number_format = '$#,##0;($#,##0);"-"'

set_col_widths(ws, {'A': 10, 'B': 12, 'C': 32, 'D': 12, 'E': 30, 'F': 14, 'G': 12, 'H': 30, 'I': 14, 'J': 12, 'K': 14, 'L': 8, 'M': 30})
ws.freeze_panes = 'A4'

# ============================================================
# SHEET 12: GENERAL_LEDGER
# ============================================================
ws = wb.create_sheet("12_GENERAL_LEDGER")
ws.sheet_properties.tabColor = TAB_ACCT
make_title(ws, "GENERAL LEDGER — Account Activity Detail", 'A1:I1')

ws.merge_cells('A3:B3')
ws['A3'].value = "Select Account Number:"
ws['A3'].font = bold_font()
ws['C3'].value = 1000
ws['C3'].font = blue_font()

dv = DataValidation(type="list", formula1='"1000,1010,1020,1030,1100,1110,1120,1300,1400,1410,1420,2000,2010,2100,2500,3000,3010,3030,4000,4020,5010,5100,5110,6100,6120,6300,6310,6800,6810"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(ws['C3'])

ws.row_dimensions[4].height = 6

hdrs = ['GL_ID', 'Date', 'Account_Number', 'Account_Name', 'Description', 'JE_Reference', 'Debit', 'Credit', 'Entity']
make_header_row(ws, 5, hdrs)

ws.cell(row=6, column=1, value='GL-001').font = data_font()
ws.cell(row=6, column=2, value='05/01/2026').font = data_font()
ws.cell(row=6, column=3, value=1000).font = data_font()
ws.cell(row=6, column=4, value='Cash — Operating Account').font = data_font()
ws.cell(row=6, column=5, value='Opening capital contribution').font = data_font()
ws.cell(row=6, column=6, value='JE001').font = data_font()
ws.cell(row=6, column=7, value=25000).font = data_font()
ws.cell(row=6, column=7).number_format = '$#,##0;($#,##0);"-"'
ws.cell(row=6, column=8, value=0).font = data_font()
ws.cell(row=6, column=8).number_format = '$#,##0;($#,##0);"-"'
ws.cell(row=6, column=9, value='STLLC').font = data_font()

for col in range(1, 10):
    ws.cell(row=6, column=col).border = thin_border()
    ws.cell(row=6, column=col).fill = white_fill()

set_col_widths(ws, {'A': 10, 'B': 12, 'C': 14, 'D': 32, 'E': 35, 'F': 14, 'G': 14, 'H': 14, 'I': 12})
ws.freeze_panes = 'A6'

# ============================================================
# SHEET 13: TRIAL_BALANCE
# ============================================================
ws = wb.create_sheet("13_TRIAL_BALANCE")
ws.sheet_properties.tabColor = TAB_ACCT
make_title(ws, "TRIAL BALANCE — As of 05/29/2026", 'A1:E1')
ws.row_dimensions[2].height = 6

hdrs = ['Account_Number', 'Account_Name', 'Account_Type', 'Debit', 'Credit']
make_header_row(ws, 3, hdrs)

# Use hardcoded values based on JE data for trial balance
tb_data = [
    (1000, 'Cash — Operating Account', 'Asset', '=SUMIF(\'11_JOURNAL_ENTRIES\'!D:D,A4,\'11_JOURNAL_ENTRIES\'!F:F)', '=SUMIF(\'11_JOURNAL_ENTRIES\'!H:H,A4,\'11_JOURNAL_ENTRIES\'!I:I)'),
    (1410, '502 Buckley — Building', 'Asset', '=SUMIF(\'11_JOURNAL_ENTRIES\'!D:D,A5,\'11_JOURNAL_ENTRIES\'!F:F)', '=SUMIF(\'11_JOURNAL_ENTRIES\'!H:H,A5,\'11_JOURNAL_ENTRIES\'!I:I)'),
    (1420, '502 Buckley — Land', 'Asset', '=SUMIF(\'11_JOURNAL_ENTRIES\'!D:D,A6,\'11_JOURNAL_ENTRIES\'!F:F)', '=SUMIF(\'11_JOURNAL_ENTRIES\'!H:H,A6,\'11_JOURNAL_ENTRIES\'!I:I)'),
    (2010, 'AP — Contractors', 'Liability', '=SUMIF(\'11_JOURNAL_ENTRIES\'!D:D,A7,\'11_JOURNAL_ENTRIES\'!F:F)', '=SUMIF(\'11_JOURNAL_ENTRIES\'!H:H,A7,\'11_JOURNAL_ENTRIES\'!I:I)'),
    (2110, 'Chase Business Card', 'Liability', '=SUMIF(\'11_JOURNAL_ENTRIES\'!D:D,A8,\'11_JOURNAL_ENTRIES\'!F:F)', '=SUMIF(\'11_JOURNAL_ENTRIES\'!H:H,A8,\'11_JOURNAL_ENTRIES\'!I:I)'),
    (2500, 'Mortgage Payable — 502 Buckley', 'Liability', '=SUMIF(\'11_JOURNAL_ENTRIES\'!D:D,A9,\'11_JOURNAL_ENTRIES\'!F:F)', '=SUMIF(\'11_JOURNAL_ENTRIES\'!H:H,A9,\'11_JOURNAL_ENTRIES\'!I:I)'),
    (3030, 'Capital Contributions', 'Equity', '=SUMIF(\'11_JOURNAL_ENTRIES\'!D:D,A10,\'11_JOURNAL_ENTRIES\'!F:F)', '=SUMIF(\'11_JOURNAL_ENTRIES\'!H:H,A10,\'11_JOURNAL_ENTRIES\'!I:I)'),
    (5010, 'Acquisition Costs', 'COGS', '=SUMIF(\'11_JOURNAL_ENTRIES\'!D:D,A11,\'11_JOURNAL_ENTRIES\'!F:F)', '=SUMIF(\'11_JOURNAL_ENTRIES\'!H:H,A11,\'11_JOURNAL_ENTRIES\'!I:I)'),
    (5100, 'Rehab — Labor', 'COGS', '=SUMIF(\'11_JOURNAL_ENTRIES\'!D:D,A12,\'11_JOURNAL_ENTRIES\'!F:F)', '=SUMIF(\'11_JOURNAL_ENTRIES\'!H:H,A12,\'11_JOURNAL_ENTRIES\'!I:I)'),
    (5110, 'Rehab — Materials', 'COGS', '=SUMIF(\'11_JOURNAL_ENTRIES\'!D:D,A13,\'11_JOURNAL_ENTRIES\'!F:F)', '=SUMIF(\'11_JOURNAL_ENTRIES\'!H:H,A13,\'11_JOURNAL_ENTRIES\'!I:I)'),
    (6120, 'Marketing — Software', 'Expense', '=SUMIF(\'11_JOURNAL_ENTRIES\'!D:D,A14,\'11_JOURNAL_ENTRIES\'!F:F)', '=SUMIF(\'11_JOURNAL_ENTRIES\'!H:H,A14,\'11_JOURNAL_ENTRIES\'!I:I)'),
]

for i, (acct, name, atype, dr, cr) in enumerate(tb_data):
    row = 4 + i
    ws.cell(row=row, column=1, value=acct).font = data_font()
    ws.cell(row=row, column=2, value=name).font = data_font()
    ws.cell(row=row, column=3, value=atype).font = data_font()
    dc = ws.cell(row=row, column=4, value=dr)
    dc.font = Font(name='Arial', size=10, color=BLACK)
    dc.number_format = '$#,##0;($#,##0);"-"'
    cc = ws.cell(row=row, column=5, value=cr)
    cc.font = Font(name='Arial', size=10, color=BLACK)
    cc.number_format = '$#,##0;($#,##0);"-"'
    style_row(ws, row, 5, alt=(i % 2 == 1))

total_row = 4 + len(tb_data) + 1
ws.merge_cells(f'A{total_row}:C{total_row}')
ws.cell(row=total_row, column=1, value='TOTAL').font = bold_font()
ws.cell(row=total_row, column=1).fill = nav_fill()
ws.cell(row=total_row, column=1).font = hdr_font(color=GOLD)

td = ws.cell(row=total_row, column=4, value=f'=SUM(D4:D{total_row-1})')
td.font = bold_font(); td.number_format = '$#,##0;($#,##0);"-"'; td.fill = nav_fill()
td.font = Font(name='Arial', size=10, bold=True, color=GOLD)

tc = ws.cell(row=total_row, column=5, value=f'=SUM(E4:E{total_row-1})')
tc.number_format = '$#,##0;($#,##0);"-"'; tc.fill = nav_fill()
tc.font = Font(name='Arial', size=10, bold=True, color=GOLD)

diff_row = total_row + 1
ws.merge_cells(f'A{diff_row}:C{diff_row}')
ws.cell(row=diff_row, column=1, value='DIFFERENCE (must = 0)').font = bold_font()
diff_c = ws.cell(row=diff_row, column=4, value=f'=D{total_row}-E{total_row}')
diff_c.font = bold_font(); diff_c.number_format = '$#,##0;($#,##0);"-"'

set_col_widths(ws, {'A': 16, 'B': 42, 'C': 16, 'D': 16, 'E': 16})
ws.freeze_panes = 'A4'

# ============================================================
# SHEET 14: BALANCE_SHEET
# ============================================================
ws = wb.create_sheet("14_BALANCE_SHEET")
ws.sheet_properties.tabColor = TAB_ACCT
make_title(ws, "BALANCE SHEET — As of 05/29/2026", 'A1:C1')

def bs_section(ws, start_row, title, items, total_label):
    ws.merge_cells(f'A{start_row}:C{start_row}')
    c = ws.cell(row=start_row, column=1, value=title)
    c.font = hdr_font(size=10); c.fill = dark_fill(); c.alignment = left()
    r = start_row + 1
    for name, val in items:
        ws.cell(row=r, column=1, value='  ' + name).font = data_font()
        vc = ws.cell(row=r, column=3, value=val)
        vc.font = blue_font() if isinstance(val, (int, float)) else Font(name='Arial', size=10, color=BLACK)
        vc.number_format = '$#,##0;($#,##0);"-"'
        vc.alignment = right_align()
        style_row(ws, r, 3, alt=(r % 2 == 0))
        r += 1
    ws.cell(row=r, column=2, value=total_label).font = bold_font()
    total_range = f'C{start_row+1}:C{r-1}'
    tc = ws.cell(row=r, column=3, value=f'=SUM({total_range})')
    tc.font = bold_font(); tc.number_format = '$#,##0;($#,##0);"-"'; tc.alignment = right_align()
    tc.border = Border(top=Side(style='thin', color='000000'), bottom=Side(style='double', color='000000'))
    return r + 1

row = 3
ws.cell(row=row, column=1, value='ASSETS').font = hdr_font(size=11, color=GOLD)
ws.cell(row=row, column=1).fill = nav_fill()
row += 1

row = bs_section(ws, row, 'CURRENT ASSETS', [
    ('Cash — Operating Account', 17200),
    ('Accounts Receivable', 0),
    ('Prepaid Expenses', 0),
], 'Total Current Assets')

row = bs_section(ws, row, 'FIXED ASSETS', [
    ('502 Buckley — Building', 97900),
    ('502 Buckley — Land', 16100),
    ('Accumulated Depreciation', 0),
], 'Total Fixed Assets')

# Total Assets
ta_row = row
ws.cell(row=ta_row, column=1, value='TOTAL ASSETS').font = bold_font(size=11)
ta_cell = ws.cell(row=ta_row, column=3, value='=C8+C13')
ta_cell.font = Font(name='Arial', size=11, bold=True, color=BLACK)
ta_cell.number_format = '$#,##0;($#,##0);"-"'
ta_cell.alignment = right_align()
ta_cell.border = Border(top=Side(style='double', color='000000'), bottom=Side(style='double', color='000000'))
row += 2

ws.cell(row=row, column=1, value='LIABILITIES & EQUITY').font = hdr_font(size=11, color=GOLD)
ws.cell(row=row, column=1).fill = nav_fill()
row += 1

row = bs_section(ws, row, 'CURRENT LIABILITIES', [
    ('AP — Contractors', 8000),
    ('Chase Business Card', 299),
    ('Accrued Expenses', 0),
], 'Total Current Liabilities')

row = bs_section(ws, row, 'LONG-TERM LIABILITIES', [
    ('Mortgage Payable — 502 Buckley', 114000),
], 'Total Long-Term Liabilities')

# Total Liabilities
tl_row = row
ws.cell(row=tl_row, column=1, value='TOTAL LIABILITIES').font = bold_font()
tl_cell = ws.cell(row=tl_row, column=3, value='=C22+C26')
tl_cell.font = bold_font(); tl_cell.number_format = '$#,##0;($#,##0);"-"'; tl_cell.alignment = right_align()
row += 1

row = bs_section(ws, row, "OWNER'S EQUITY", [
    ('Capital Contributions', 25000),
    ('Members Draws', 0),
    ('Retained Earnings', -8799),
], "Total Owner's Equity")

# Total L+E
tle_row = row
ws.cell(row=tle_row, column=1, value='TOTAL LIABILITIES + EQUITY').font = bold_font(size=11)
tle_cell = ws.cell(row=tle_row, column=3, value=f'=C{tl_row}+C{tle_row-2}')
tle_cell.font = Font(name='Arial', size=11, bold=True, color=BLACK)
tle_cell.number_format = '$#,##0;($#,##0);"-"'; tle_cell.alignment = right_align()
row += 2

# Balance Check
ws.cell(row=row, column=1, value='BALANCE CHECK (must = $0):').font = bold_font()
bc = ws.cell(row=row, column=3, value=f'=C{ta_row}-C{tle_row}')
bc.font = bold_font(); bc.number_format = '$#,##0;($#,##0);"-"'; bc.alignment = right_align()

set_col_widths(ws, {'A': 35, 'B': 25, 'C': 18})
ws.freeze_panes = 'A3'

# ============================================================
# SHEET 15: INCOME_STATEMENT
# ============================================================
ws = wb.create_sheet("15_INCOME_STATEMENT")
ws.sheet_properties.tabColor = TAB_ACCT
make_title(ws, "INCOME STATEMENT — Period: Jan 2026 — May 2026", 'A1:C1')

def is_section(ws, row, title, items):
    ws.merge_cells(f'A{row}:C{row}')
    c = ws.cell(row=row, column=1, value=title)
    c.font = hdr_font(size=10); c.fill = dark_fill(); c.alignment = left()
    r = row + 1
    for name, val in items:
        ws.cell(row=r, column=1, value='  ' + name).font = data_font()
        vc = ws.cell(row=r, column=3, value=val)
        vc.font = Font(name='Arial', size=10, color=BLACK)
        vc.number_format = '$#,##0;($#,##0);"-"'; vc.alignment = right_align()
        style_row(ws, r, 3, alt=(r % 2 == 0))
        r += 1
    return r

row = 3
row = is_section(ws, row, 'REVENUE', [
    ('Wholesale Assignment Fees', 0),
    ('Flip Profits', 0),
    ('Rental Income', 0),
    ('Other Revenue', 0),
])
rev_end = row - 1
rev_total_row = row
ws.cell(row=rev_total_row, column=2, value='TOTAL REVENUE').font = bold_font()
ws.cell(row=rev_total_row, column=3, value=f'=SUM(C4:C{rev_end})').font = bold_font()
ws.cell(row=rev_total_row, column=3).number_format = '$#,##0;($#,##0);"-"'
ws.cell(row=rev_total_row, column=3).alignment = right_align()
row += 2

row = is_section(ws, row, 'COST OF GOODS SOLD (COGS)', [
    ('Acquisition Costs', 4500),
    ('Rehab — Labor', 8000),
    ('Rehab — Materials', 3500),
    ('Holding Costs', 1080),
    ('Assignment Costs', 0),
])
cogs_end = row - 1
cogs_total_row = row
ws.cell(row=cogs_total_row, column=2, value='TOTAL COGS').font = bold_font()
ws.cell(row=cogs_total_row, column=3, value=f'=SUM(C{rev_total_row+2}:C{cogs_end})').font = bold_font()
ws.cell(row=cogs_total_row, column=3).number_format = '$#,##0;($#,##0);"-"'
ws.cell(row=cogs_total_row, column=3).alignment = right_align()
row += 1

gp_row = row
ws.cell(row=gp_row, column=2, value='GROSS PROFIT').font = bold_font(size=11)
gp = ws.cell(row=gp_row, column=3, value=f'=C{rev_total_row}-C{cogs_total_row}')
gp.font = bold_font(size=11); gp.number_format = '$#,##0;($#,##0);"-"'; gp.alignment = right_align()
row += 1
ws.cell(row=row, column=2, value='Gross Margin %').font = data_font()
gm = ws.cell(row=row, column=3, value=f'=IF(C{rev_total_row}=0,0,C{gp_row}/C{rev_total_row})')
gm.font = Font(name='Arial', size=10, color=BLACK); gm.number_format = '0.0%'; gm.alignment = right_align()
row += 2

row = is_section(ws, row, 'OPERATING EXPENSES', [
    ('Marketing', 299),
    ('Professional Fees', 0),
    ('Software & Technology', 0),
    ('Insurance', 0),
    ('Vehicle & Travel', 0),
    ('Bank Fees', 0),
    ('Other Expenses', 0),
])
opex_end = row - 1
opex_total_row = row
ws.cell(row=opex_total_row, column=2, value='TOTAL OPERATING EXPENSES').font = bold_font()
ws.cell(row=opex_total_row, column=3, value=f'=SUM(C{opex_total_row-8}:C{opex_end})').font = bold_font()
ws.cell(row=opex_total_row, column=3).number_format = '$#,##0;($#,##0);"-"'
ws.cell(row=opex_total_row, column=3).alignment = right_align()
row += 1

ebit_row = row
ws.cell(row=ebit_row, column=2, value='OPERATING INCOME (EBIT)').font = bold_font(size=11)
ebit = ws.cell(row=ebit_row, column=3, value=f'=C{gp_row}-C{opex_total_row}')
ebit.font = bold_font(size=11); ebit.number_format = '$#,##0;($#,##0);"-"'; ebit.alignment = right_align()
row += 2

row = is_section(ws, row, 'BELOW-THE-LINE', [
    ('Interest Expense', 0),
    ('Depreciation', 0),
])
bel_end = row - 1
row += 1

ni_row = row
ws.cell(row=ni_row, column=2, value='NET INCOME').font = bold_font(size=12)
ni = ws.cell(row=ni_row, column=3, value=f'=C{ebit_row}-SUM(C{ebit_row+2}:C{bel_end})')
ni.font = Font(name='Arial', size=12, bold=True, color=BLACK)
ni.number_format = '$#,##0;($#,##0);"-"'; ni.alignment = right_align()
ni.border = Border(top=Side(style='double', color='000000'), bottom=Side(style='double', color='000000'))

set_col_widths(ws, {'A': 35, 'B': 30, 'C': 18})
ws.freeze_panes = 'A3'

# ============================================================
# SHEET 16: CASH_FLOW
# ============================================================
ws = wb.create_sheet("16_CASH_FLOW")
ws.sheet_properties.tabColor = TAB_ACCT
make_title(ws, "CASH FLOW STATEMENT — YTD 2026", 'A1:C1')

def cf_section(ws, start_row, title, items, total_label):
    ws.merge_cells(f'A{start_row}:C{start_row}')
    c = ws.cell(row=start_row, column=1, value=title)
    c.font = hdr_font(size=10); c.fill = dark_fill(); c.alignment = left()
    r = start_row + 1
    for name, val in items:
        ws.cell(row=r, column=1, value='  ' + name).font = data_font()
        vc = ws.cell(row=r, column=3, value=val)
        vc.font = blue_font() if isinstance(val, (int, float)) else Font(name='Arial', size=10, color=BLACK)
        vc.number_format = '$#,##0;($#,##0);"-"'; vc.alignment = right_align()
        style_row(ws, r, 3, alt=(r % 2 == 0))
        r += 1
    ws.cell(row=r, column=2, value=total_label).font = bold_font()
    tc = ws.cell(row=r, column=3, value=f'=SUM(C{start_row+1}:C{r-1})')
    tc.font = bold_font(); tc.number_format = '$#,##0;($#,##0);"-"'; tc.alignment = right_align()
    return r + 1

row = 3
row = cf_section(ws, row, 'A. OPERATING ACTIVITIES', [
    ('Net Income', -16379),
    ('Add: Depreciation', 0),
    ('Changes in Accounts Receivable', 0),
    ('Changes in Accounts Payable', 8299),
], 'Net Operating Cash Flow')
op_total = row - 1
row += 1

row = cf_section(ws, row, 'B. INVESTING ACTIVITIES', [
    ('Property Acquisitions', -114000),
    ('Rehab Spending', -11500),
    ('Proceeds from Property Sales', 0),
], 'Net Investing Cash Flow')
inv_total = row - 1
row += 1

row = cf_section(ws, row, 'C. FINANCING ACTIVITIES', [
    ('Mortgage Proceeds', 114000),
    ('Mortgage Payments', 0),
    ('Capital Contributions', 25000),
    ('Owner Draws', 0),
], 'Net Financing Cash Flow')
fin_total = row - 1
row += 2

ws.cell(row=row, column=1, value='NET CHANGE IN CASH').font = bold_font(size=11)
nc = ws.cell(row=row, column=3, value=f'=C{op_total}+C{inv_total}+C{fin_total}')
nc.font = bold_font(size=11); nc.number_format = '$#,##0;($#,##0);"-"'; nc.alignment = right_align()
row += 1

ws.cell(row=row, column=1, value='Beginning Cash Balance').font = data_font()
bc = ws.cell(row=row, column=3, value=0)
bc.font = blue_font(); bc.number_format = '$#,##0;($#,##0);"-"'; bc.alignment = right_align()
row += 1

ws.cell(row=row, column=1, value='ENDING CASH BALANCE').font = bold_font(size=11)
ec = ws.cell(row=row, column=3, value=f'=C{row-2}+C{row-1}')
ec.font = Font(name='Arial', size=11, bold=True, color=BLACK)
ec.number_format = '$#,##0;($#,##0);"-"'; ec.alignment = right_align()
ec.border = Border(top=Side(style='double', color='000000'), bottom=Side(style='double', color='000000'))

set_col_widths(ws, {'A': 38, 'B': 28, 'C': 18})
ws.freeze_panes = 'A3'

# ============================================================
# SHEET 17: BUDGETS
# ============================================================
ws = wb.create_sheet("17_BUDGETS")
ws.sheet_properties.tabColor = TAB_ACCT
make_title(ws, "ANNUAL BUDGET — FY 2026", 'A1:P1')
ws.row_dimensions[2].height = 6

months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
hdrs = ['Account_Number', 'Account_Name', 'Category'] + months + ['Annual_Total']
make_header_row(ws, 3, hdrs)

budget_data = [
    (4000, 'Wholesale Assignment Fees', 'Revenue', [5000]*12),
    (4020, 'Rental Income', 'Revenue', [0]*12),
    (5100, 'Rehab — Labor', 'COGS', [8000,8000,5000,3000,2000,0,0,0,0,0,0,0]),
    (5110, 'Rehab — Materials', 'COGS', [5000,5000,3000,2000,1000,0,0,0,0,0,0,0]),
    (5210, 'Holding — Mortgage', 'COGS', [750]*12),
    (5220, 'Holding — Insurance', 'COGS', [100]*12),
    (6100, 'Marketing — Digital', 'Expense', [1500]*12),
    (6120, 'Marketing — Software', 'Expense', [299]*12),
    (6300, 'Professional Fees — Legal', 'Expense', [500,0,500,0,500,0,500,0,500,0,500,0]),
    (6310, 'Professional Fees — Accounting', 'Expense', [250]*12),
    (6400, 'Software & Technology', 'Expense', [200]*12),
    (6500, 'Insurance', 'Expense', [150]*12),
    (6600, 'Vehicle & Travel', 'Expense', [300]*12),
    (6700, 'Bank Fees', 'Expense', [25]*12),
]

for i, (acct, name, cat, monthly) in enumerate(budget_data):
    row = 4 + i
    ws.cell(row=row, column=1, value=acct).font = data_font()
    ws.cell(row=row, column=2, value=name).font = data_font()
    ws.cell(row=row, column=3, value=cat).font = data_font()
    for m, val in enumerate(monthly):
        c = ws.cell(row=row, column=4+m, value=val)
        c.font = blue_font(); c.number_format = '$#,##0;($#,##0);"-"'; c.alignment = right_align()
    tc = ws.cell(row=row, column=16, value=f'=SUM(D{row}:O{row})')
    tc.font = Font(name='Arial', size=10, bold=True, color=BLACK)
    tc.number_format = '$#,##0;($#,##0);"-"'; tc.alignment = right_align()
    style_row(ws, row, 16, alt=(i % 2 == 1))

set_col_widths(ws, {'A': 14, 'B': 38, 'C': 12, 'D': 10, 'E': 10, 'F': 10, 'G': 10, 'H': 10, 'I': 10, 'J': 10, 'K': 10, 'L': 10, 'M': 10, 'N': 10, 'O': 10, 'P': 14})
ws.freeze_panes = 'D4'

# ============================================================
# SHEET 18: BUDGET_VS_ACTUAL
# ============================================================
ws = wb.create_sheet("18_BUDGET_VS_ACTUAL")
ws.sheet_properties.tabColor = TAB_ACCT
make_title(ws, "BUDGET VS ACTUAL — YTD 2026", 'A1:G1')
ws.row_dimensions[2].height = 6

hdrs = ['Account_Number', 'Account_Name', 'Budget_YTD', 'Actual_YTD', 'Variance', 'Variance_%', 'Status']
make_header_row(ws, 3, hdrs)

bva_data = [
    (4000, 'Wholesale Assignment Fees'),
    (5100, 'Rehab — Labor'),
    (5110, 'Rehab — Materials'),
    (5210, 'Holding — Mortgage'),
    (6100, 'Marketing — Digital'),
    (6120, 'Marketing — Software'),
    (6300, 'Professional Fees — Legal'),
    (6310, 'Professional Fees — Accounting'),
]

budget_ytd = [25000, 26000, 16000, 3750, 7500, 1495, 1500, 1250]
actual_ytd = [0, 8000, 3500, 750, 0, 299, 0, 0]

for i, ((acct, name), bud, act) in enumerate(zip(bva_data, budget_ytd, actual_ytd)):
    row = 4 + i
    ws.cell(row=row, column=1, value=acct).font = data_font()
    ws.cell(row=row, column=2, value=name).font = data_font()
    bc = ws.cell(row=row, column=3, value=bud)
    bc.font = blue_font(); bc.number_format = '$#,##0;($#,##0);"-"'; bc.alignment = right_align()
    ac = ws.cell(row=row, column=4, value=act)
    ac.font = blue_font(); ac.number_format = '$#,##0;($#,##0);"-"'; ac.alignment = right_align()
    vc = ws.cell(row=row, column=5, value=f'=C{row}-D{row}')
    vc.font = Font(name='Arial', size=10, color=BLACK); vc.number_format = '$#,##0;($#,##0);"-"'; vc.alignment = right_align()
    vp = ws.cell(row=row, column=6, value=f'=IF(C{row}=0,0,(C{row}-D{row})/C{row})')
    vp.font = Font(name='Arial', size=10, color=BLACK); vp.number_format = '0.0%'; vp.alignment = right_align()
    st = ws.cell(row=row, column=7, value=f'=IF(C{row}=0,"N/A",IF(E{row}<0,"OVER BUDGET",IF(ABS(F{row})<0.1,"ON TRACK","UNDER BUDGET")))')
    st.font = Font(name='Arial', size=10, color=BLACK)
    style_row(ws, row, 7, alt=(i % 2 == 1))

set_col_widths(ws, {'A': 14, 'B': 38, 'C': 14, 'D': 14, 'E': 14, 'F': 12, 'G': 16})
ws.freeze_panes = 'A4'

wb.save('C:/stllcweb3/Dynasty_Accounting_OS_v1.xlsx')
print("Accounting Core sheets (10-18) done")
